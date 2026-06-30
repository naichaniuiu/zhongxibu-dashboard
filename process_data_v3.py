import openpyxl
import json
import os
from datetime import datetime, timedelta
from collections import defaultdict, Counter

# ============================================================
# 中西部大区 26 财年 Q1 数据看板 - 新数据源整合处理
# 数据源：
#   1) 业绩数据：D:/26财年Q1业绩数据.xlsx
#   2) 认款数据：D:/认款数据.xlsx
#   3) 欠款数据：D:/欠款数据.xlsx
#   4) 25Q1 数据：D:/25财年Q1数据.xlsx
# ============================================================

TODAY = datetime(2026, 6, 30)  # 数据基准日
Q1_START = datetime(2026, 4, 1)
Q1_END = datetime(2026, 6, 30)
Q1_START_25 = datetime(2025, 4, 1)
Q1_END_25 = datetime(2025, 6, 30)
TARGET_TOTAL = 5586.0  # 万元，沿用旧看板目标
BLACKLIST = {'支振岗', '李国栋', '白雨'}
KEEP_LIST = {'张宸睿'}

DEPT_MAP = {
    '中西部大区': '中西部大区',
    '华中大区': '中西部大区',
    '华中大区（已封存）': '中西部大区',
    '西南大区': '中西部大区',
    '西南大区（已封存）': '中西部大区',
}


# ============================================================
# 部门映射规则（二级部门 & 三级部门）
# ============================================================
INTERNET_SELLERS = set()


def scan_internet_sellers(path):
    """扫描文件中原始部门为武汉通讯互联网的销售员，用于后续按销售员归属映射。"""
    for r in load_rows(path):
        dept1 = str(r.get('一级部门') or '').strip().replace('\t', '')
        if dept1 not in DEPT_MAP:
            continue
        dept2 = str(r.get('二级部门') or '').strip().replace('\t', '')
        sub_dept = str(r.get('三级部门') or '').strip().replace('\t', '')
        if '通讯互联网' in dept2 or '通讯互联网' in sub_dept:
            seller = str(r.get('销售员名称') or '').strip().replace('\t', '')
            if seller:
                INTERNET_SELLERS.add(seller)


def normalize_dept2(dept2, sub_dept, seller_name):
    """根据二级部门、三级部门和销售员名称归一化二级部门。"""
    dept2 = str(dept2 or '其他').strip().replace('\t', '')
    sub_dept = str(sub_dept or '').strip().replace('\t', '')
    seller = str(seller_name or '').strip().replace('\t', '')

    # 武汉通讯互联网：先按销售员归属拆分（适用于该销售员的所有记录）
    if seller in INTERNET_SELLERS:
        if seller in ('吴晗', '李国栋'):
            return '混营销区'
        return '湖北营销区'
    # 再按原始部门名称匹配（兼容数据未归集到销售员的情况）
    if '通讯互联网' in dept2 or '通讯互联网' in sub_dept:
        if seller in ('吴晗', '李国栋'):
            return '混营销区'
        return '湖北营销区'

    # 武汉金融 / 武汉能源交通 / 武汉基建制造 → 湖北营销区
    if dept2 in ('武汉金融', '武汉能源交通', '武汉基建制造') or \
       sub_dept in ('武汉金融行业组', '武汉能源交通行业组', '武汉基建制造行业组'):
        return '湖北营销区'

    # 四川 / 重庆营销区 → 综合管理办公室
    if dept2 in ('四川营销区', '重庆营销区'):
        return '综合管理办公室'

    # 其他城市部门：成都 / 重庆 / 郑州 / 长沙 / 西安
    if dept2 in ('成都', '重庆', '郑州', '长沙', '西安') or \
       sub_dept in ('成都站', '重庆站', '郑州站', '长沙站', '西安站'):
        return '综合管理办公室'

    return dept2


def normalize_sub_dept(dept2, sub_dept):
    """根据二级部门归一化三级部门。"""
    dept2 = str(dept2 or '').strip().replace('\t', '')
    sub_dept = str(sub_dept or '其他').strip().replace('\t', '')

    if dept2 == '混营销区':
        return '其他'

    if '成都' in sub_dept:
        return '成都站'
    if '重庆' in sub_dept:
        return '重庆站'

    return sub_dept


def parse_date(val):
    """解析日期字符串，去除制表符和时间部分"""
    if not val:
        return None
    s = str(val).strip().replace('\t', '').split(' ')[0]
    if not s or s == 'None':
        return None
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def to_wan(val):
    """金额元 -> 万元"""
    if val is None:
        return 0.0
    try:
        return float(val) / 10000.0
    except (TypeError, ValueError):
        return 0.0


def load_rows(path, sheet_idx=0):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[sheet_idx]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield dict(zip(headers, row))


def weighted_avg(items):
    if not items:
        return 0.0
    total_w = sum(x['amount'] for x in items)
    if total_w <= 0:
        return 0.0
    return sum(x['days'] * x['amount'] for x in items) / total_w


# ============================================================
# 0. 预扫描：识别武汉通讯互联网部门的销售员
# ============================================================
print('Scanning internet sales sellers...')
scan_internet_sellers('D:/26财年Q1业绩数据.xlsx')
scan_internet_sellers('D:/欠款数据.xlsx')
print(f'  Internet sellers: {len(INTERNET_SELLERS)}')

# ============================================================
# 1. 读取业绩数据（26 财年 Q1）
# ============================================================
print('Loading performance data...')
perf_records = []
for r in load_rows('D:/26财年Q1业绩数据.xlsx'):
    dept1 = str(r.get('一级部门') or '').strip().replace('\t', '')
    if dept1 not in DEPT_MAP:
        continue
    d = parse_date(r.get('业绩日期'))
    if not d or not (Q1_START <= d <= Q1_END):
        continue
    is_b = str(r.get('是否核算B端业绩') or '').strip().replace('\t', '')
    if is_b == '否':
        continue
    raw_dept2 = str(r.get('二级部门') or '').strip().replace('\t', '')
    raw_sub_dept = str(r.get('三级部门') or '').strip().replace('\t', '')
    seller_name = str(r.get('销售员名称') or '').strip().replace('\t', '')
    dept2 = normalize_dept2(raw_dept2, raw_sub_dept, seller_name)
    sub_dept = normalize_sub_dept(dept2, raw_sub_dept)
    perf_records.append({
        'date': d,
        'order_no': str(r.get('业绩单号') or '').strip().replace('\t', ''),
        'customer_id': str(r.get('客户编号') or '').strip().replace('\t', ''),
        'customer_name': str(r.get('客户名称') or '').strip().replace('\t', ''),
        'seller_no': str(r.get('销售员工号') or '').strip().replace('\t', ''),
        'seller_name': seller_name,
        'dept': dept2,
        'sub_dept': sub_dept,
        'seller_status': str(r.get('销售员状态') or '').strip().replace('\t', ''),
        'perf': to_wan(r.get('业绩总金额')),
        'collect': to_wan(r.get('回款金额')),
        'debt': to_wan(r.get('欠款金额')),
    })
print(f'  Performance records: {len(perf_records)}')

# ============================================================
# 2. 读取 25 财年 Q1 业绩数据（用于同比）
# ============================================================
print('Loading 25Q1 performance data...')
perf_records_25 = []
for r in load_rows('D:/25财年Q1数据.xlsx'):
    dept1 = str(r.get('一级部门') or '').strip().replace('\t', '')
    if dept1 not in DEPT_MAP:
        continue
    d = parse_date(r.get('业绩日期'))
    if not d or not (Q1_START_25 <= d <= Q1_END_25):
        continue
    is_b = str(r.get('是否核算B端业绩') or '').strip().replace('\t', '')
    if is_b == '否':
        continue
    raw_dept2 = str(r.get('二级部门') or '').strip().replace('\t', '')
    raw_sub_dept = str(r.get('三级部门') or '').strip().replace('\t', '')
    seller_name = str(r.get('销售员名称') or '').strip().replace('\t', '')
    dept2 = normalize_dept2(raw_dept2, raw_sub_dept, seller_name)
    perf_records_25.append({
        'dept': dept2,
        'seller_name': seller_name,
        'perf': to_wan(r.get('业绩总金额')),
    })
print(f'  25Q1 performance records: {len(perf_records_25)}')

# ============================================================
# 3. 读取欠款数据（整个文件所有行直接求和，无过滤）
# ============================================================
print('Loading debt data...')
# 按行直接处理，不按订单聚合，不过滤部门和日期
debt_records = []
for r in load_rows('D:/欠款数据.xlsx'):
    d = parse_date(r.get('业绩日期'))
    raw_dept2 = str(r.get('二级部门') or '').strip().replace('\t', '')
    raw_sub_dept = str(r.get('三级部门') or '').strip().replace('\t', '')
    seller_name = str(r.get('销售员名称') or '').strip().replace('\t', '')
    dept2 = normalize_dept2(raw_dept2, raw_sub_dept, seller_name)
    order_no = str(r.get('业绩单号') or '').strip().replace('\t', '')
    debt_val = to_wan(r.get('欠款金额'))
    days = (TODAY - d).days if d else 0
    if days < 0:
        days = 0
    debt_records.append({
        'date': d,
        'order_no': order_no,
        'customer_id': str(r.get('客户编号') or '').strip().replace('\t', ''),
        'customer_name': str(r.get('客户名称') or '').strip().replace('\t', ''),
        'seller_no': str(r.get('销售员工号') or '').strip().replace('\t', ''),
        'seller_name': seller_name,
        'seller_status': str(r.get('销售员状态') or '').strip().replace('\t', ''),
        'dept': dept2,
        'debt': debt_val,
        'days': days,
    })

print(f'  Debt rows: {len(debt_records)}, total debt: {sum(r["debt"] for r in debt_records):.2f}')

# ============================================================
# 4. 读取认款数据，计算回款周期
# ============================================================
print('Loading payment data...')
# 建立业绩单号 -> 部门 映射
order_to_dept = {}
for rec in perf_records:
    if rec['order_no']:
        order_to_dept.setdefault(rec['order_no'], rec['dept'])

# 回款周期记录：{dept, seller, days, amount}
cycle_records = []
for r in load_rows('D:/认款数据.xlsx'):
    if str(r.get('目标认款类型') or '').strip() != '业绩单认款':
        continue
    order_no = str(r.get('业绩单号') or '').strip().replace('\t', '')
    dept = order_to_dept.get(order_no)
    if not dept:
        continue
    is_b = str(r.get('是否核算B端业绩') or '').strip().replace('\t', '')
    if is_b == '否':
        continue
    perf_date = parse_date(r.get('业绩日期'))
    pay_date = parse_date(r.get('回款日期'))
    if not perf_date or not pay_date:
        continue
    days = (pay_date - perf_date).days
    if days < 0 or days > 365 * 2:
        continue
    amount = to_wan(r.get('认款协同金额'))
    seller = str(r.get('销售员名称') or '').strip().replace('\t', '')
    cycle_records.append({'dept': dept, 'seller': seller, 'days': days, 'amount': amount})

# 按部门聚合回款周期
dept_cycle_items = defaultdict(list)
seller_cycle_items = defaultdict(list)
for rec in cycle_records:
    dept_cycle_items[rec['dept']].append(rec)
    if rec['seller']:
        seller_cycle_items[rec['seller']].append(rec)

avg_cycle = weighted_avg(cycle_records)
print(f'  Payment records for cycle: {len(cycle_records)}, avg cycle: {avg_cycle:.1f}')

# ============================================================
# 5. 销售员维度聚合（用于弹窗下钻）
# ============================================================
print('Aggregating by seller...')

seller_data = defaultdict(lambda: {
    'perf': 0.0, 'collect': 0.0, 'total_debt': 0.0,
    'd30': 0.0, 'd30_90': 0.0, 'd90_180': 0.0, 'd180': 0.0,
    'dept': '其他',
})

for r in perf_records:
    s = seller_data[r['seller_name']]
    s['perf'] += r['perf']
    s['collect'] += r['collect']
    s['dept'] = r['dept']

for r in debt_records:
    s = seller_data[r['seller_name']]
    s['total_debt'] += r['debt']
    s['dept'] = r['dept']
    if r['days'] <= 30:
        s['d30'] += r['debt']
    elif r['days'] <= 90:
        s['d30_90'] += r['debt']
    elif r['days'] <= 180:
        s['d90_180'] += r['debt']
    else:
        s['d180'] += r['debt']

# 按部门聚合销售员（用于弹窗）
sales_detail_data = defaultdict(list)
sales_cycle_detail = defaultdict(list)
for seller, s in seller_data.items():
    dept = s['dept']
    sales_detail_data[dept].append({
        'name': seller,
        'perf': round(s['perf'], 2),
        'collect': round(s['collect'], 2),
        'total_debt': round(s['total_debt'], 2),
        'd30': round(s['d30'], 2),
        'd30_90': round(s['d30_90'], 2),
        'd90_180': round(s['d90_180'], 2),
        'd180': round(s['d180'], 2),
    })
    c_items = seller_cycle_items.get(seller, [])
    cycle = weighted_avg(c_items)
    sales_cycle_detail[dept].append({
        'name': seller,
        'debt_amt': round(s['total_debt'] * 10000, 2),
        'rec_amt': round(s['collect'] * 10000, 2),
        'cycle': round(cycle, 1),
    })

# 转为普通 dict
sales_detail_data = dict(sales_detail_data)
sales_cycle_detail = dict(sales_cycle_detail)

# ============================================================
# 6. 计算 KPI
# ============================================================
print('Calculating KPIs...')

total_perf = sum(r['perf'] for r in perf_records)
total_collect = sum(r['collect'] for r in perf_records)
total_debt = sum(r['debt'] for r in debt_records)

total_completion = round(total_perf / TARGET_TOTAL * 100, 1) if TARGET_TOTAL > 0 else 0.0

# 在职销售员：26Q1 有实际业绩（>0）的在职销售员，剔除黑名单
seller_perf_q1 = defaultdict(float)
for r in perf_records:
    if r['seller_status'] == '在职' and r['perf'] > 0:
        seller_perf_q1[r['seller_name']] += r['perf']

active_sellers = (set(seller_perf_q1.keys()) - BLACKLIST) | KEEP_LIST
active_seller_count = len(active_sellers)

# 25Q1 总业绩
total_perf_25 = sum(r['perf'] for r in perf_records_25)
total_yoy = round((total_perf - total_perf_25) / total_perf_25 * 100, 1) if total_perf_25 > 0 else None

# 账龄分布（2026 年欠款）
d30 = sum(r['debt'] for r in debt_records if r['days'] <= 30)
d30_90 = sum(r['debt'] for r in debt_records if 30 < r['days'] <= 90)
d90_180 = sum(r['debt'] for r in debt_records if 90 < r['days'] <= 180)
d180 = sum(r['debt'] for r in debt_records if r['days'] > 180)
# 逾期欠款：>30 天
d_overdue = d30_90 + d90_180 + d180
debt_d90_plus = d90_180 + d180

# 高风险客户：欠款 > 50 万 或 账龄 > 90 天
high_risk_customers = []
for r in debt_records:
    if r['debt'] >= 50.0 or r['days'] > 90:
        high_risk_customers.append({
            'customer': r['customer_name'],
            'debt': round(r['debt'], 2),
            'days': r['days'],
            'seller': r['seller_name'],
        })
high_risk_customers.sort(key=lambda x: (-x['debt'], -x['days']))
high_risk_customers = high_risk_customers[:50]

# ============================================================
# 7. 按部门（二级部门）聚合
# ============================================================
print('Aggregating by department...')

# 26Q1 部门业绩
dept_perf = defaultdict(lambda: {'perf': 0.0, 'collect': 0.0, 'sales': set(), 'orders': set()})
for r in perf_records:
    d = dept_perf[r['dept']]
    d['perf'] += r['perf']
    d['collect'] += r['collect']
    d['sales'].add(r['seller_name'])
    d['orders'].add(r['order_no'])

# 25Q1 部门业绩
dept_perf_25 = defaultdict(float)
for r in perf_records_25:
    dept_perf_25[r['dept']] += r['perf']

# 2026 年部门欠款
dept_debt = defaultdict(lambda: {'d30': 0.0, 'd30_90': 0.0, 'd90_180': 0.0, 'd180': 0.0, 'total_debt': 0.0})
for r in debt_records:
    d = dept_debt[r['dept']]
    d['total_debt'] += r['debt']
    if r['days'] <= 30:
        d['d30'] += r['debt']
    elif r['days'] <= 90:
        d['d30_90'] += r['debt']
    elif r['days'] <= 180:
        d['d90_180'] += r['debt']
    else:
        d['d180'] += r['debt']

# 按当前业绩比例分配目标
target_ratio = {}
if total_perf > 0:
    for dept, v in dept_perf.items():
        target_ratio[dept] = v['perf'] / total_perf
else:
    for dept in dept_perf:
        target_ratio[dept] = 0.0

# 合并部门数据
all_depts = set(dept_perf.keys()) | set(dept_debt.keys()) | set(dept_perf_25.keys())
dept_data = []
for dept in sorted(all_depts):
    p = dept_perf.get(dept, {'perf': 0.0, 'collect': 0.0, 'sales': set(), 'orders': set()})
    d = dept_debt.get(dept, {'d30': 0.0, 'd30_90': 0.0, 'd90_180': 0.0, 'd180': 0.0, 'total_debt': 0.0})
    v25 = dept_perf_25.get(dept, 0.0)
    target = TARGET_TOTAL * target_ratio.get(dept, 0.0)
    completion = round(p['perf'] / target * 100, 1) if target > 0 else 0.0
    yoy = round((p['perf'] - v25) / v25 * 100, 1) if v25 > 0 else None
    cycle = weighted_avg(dept_cycle_items.get(dept, []))
    dept_data.append({
        'dept': dept,
        'v26': round(p['perf'], 2),
        'v25': round(v25, 2) if v25 > 0 else None,
        'yoy': yoy,
        'target': round(target, 2),
        'completion': completion,
        'sales': len(p['sales']),
        'd30': round(d['d30'], 2),
        'd30_90': round(d['d30_90'], 2),
        'd90_180': round(d['d90_180'], 2),
        'd180': round(d['d180'], 2),
        'total_debt': round(d['total_debt'], 2),
        'collect': round(p['collect'], 2),
        'cycle': round(cycle, 1),
    })

# 排序：按业绩金额降序
dept_data.sort(key=lambda x: -x['v26'])

# 回款周期极值
cycles_with_data = [d for d in dept_data if d['cycle'] > 0]
if cycles_with_data:
    max_cycle_dept = max(cycles_with_data, key=lambda x: x['cycle'])['dept']
    max_cycle = max(d['cycle'] for d in cycles_with_data)
    min_cycle_dept = min(cycles_with_data, key=lambda x: x['cycle'])['dept']
    min_cycle = min(d['cycle'] for d in cycles_with_data)
else:
    max_cycle_dept = ''
    max_cycle = 0.0
    min_cycle_dept = ''
    min_cycle = 0.0
over90_depts = [d['dept'] for d in cycles_with_data if d['cycle'] > 90]

# ============================================================
# 8. 输出 dashboard_data.json
# ============================================================
print('Writing dashboard_data.json...')

# 总回款比例
d30_ratio = round(d30 / total_debt * 100, 1) if total_debt > 0 else 0.0

dashboard_data = {
    'data_date': Q1_END.strftime('%Y-%m-%d'),
    'today': TODAY.strftime('%Y-%m-%d'),
    'source_files': {
        'performance': '26财年Q1业绩数据.xlsx',
        'payment': '认款数据.xlsx',
        'debt': '欠款数据.xlsx',
        'perf25': '25财年Q1数据.xlsx',
    },
    'kpi': {
        'total_perf26': round(total_perf, 2),
        'total_target': TARGET_TOTAL,
        'total_completion': total_completion,
        'total_yoy': total_yoy,
        'total_perf25': round(total_perf_25, 2),
        'total_active_sellers': active_seller_count,
        'total_debt': round(total_debt, 2),
        'debt_d90_plus': round(debt_d90_plus, 2),
        'total_overdue': round(d_overdue, 2),
    },
    'debt_kpi': {
        'total': round(total_debt, 2),
        'd30': round(d30, 2),
        'd30_ratio': d30_ratio,
        'd30_90': round(d30_90, 2),
        'd30_90_ratio': round(d30_90 / total_debt * 100, 1) if total_debt > 0 else 0.0,
        'd90_180': round(d90_180, 2),
        'd180': round(d180, 2),
        'overdue': round(d_overdue, 2),
    },
    'cycle_kpi': {
        'avg_cycle': round(avg_cycle, 1),
        'max_cycle_dept': max_cycle_dept,
        'max_cycle': round(max_cycle, 1),
        'min_cycle_dept': min_cycle_dept,
        'min_cycle': round(min_cycle, 1),
        'over90_depts': over90_depts,
        'over90_count': len(over90_depts),
    },
    'dept_data': dept_data,
    'sales_detail_data': sales_detail_data,
    'sales_cycle_detail': sales_cycle_detail,
    'high_risk_customers': high_risk_customers,
}

base_dir = os.path.dirname(os.path.abspath(__file__))
with open(os.path.join(base_dir, 'dashboard_data.json'), 'w', encoding='utf-8') as f:
    json.dump(dashboard_data, f, ensure_ascii=False, indent=2)

print('Done.')
print(f'KPI: 业绩 {total_perf:.2f} 万, 25Q1 {total_perf_25:.2f} 万, 同比 {total_yoy}%, 目标 {TARGET_TOTAL} 万, 完成率 {total_completion}%, 在职 {active_seller_count} 人, 欠款 {total_debt:.2f} 万, 逾期 {d_overdue:.2f} 万')
