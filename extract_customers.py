# -*- coding: utf-8 -*-
"""
提取客户维度完整数据，生成三级下钻结构：
部门 -> 销售员 -> 客户

业绩Tab: 部门->销售员->客户（业绩额、回款额、欠款额）
欠款Tab: 部门->销售员->客户（欠款分布、天数）
回款周期Tab: 部门->销售员->客户（回款周期）
"""
import json
import os
import sys
from collections import defaultdict
from datetime import datetime
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(os.path.expanduser('~'), 'Downloads', '业绩 欠款看板.xlsx')

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

def read_sheet(sname):
    ws = wb[sname]
    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c is not None else '' for c in row]
            continue
        if all(c is None for c in row):
            continue
        row_dict = {}
        for j, val in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = val
        rows.append(row_dict)
    return rows

# 读各sheet
rows_26 = read_sheet('26财年Q1业绩')
rows_25 = read_sheet('25财年Q1业绩')
rows_debt = read_sheet('欠款数据 ')

# 部门映射
DEPT_MAP = {
    '武汉基建制造行业组': '武汉基建制造行业组',
    '武汉金融行业组': '武汉金融行业组',
    '武汉能源交通行业组': '武汉能源交通行业组',
    '成都站': '成都站',
    '长沙站': '长沙站',
    '西安站': '西安站',
    '郑州站': '郑州站',
    '重庆站': '重庆站',
}

def get_dept(row):
    d3 = str(row.get('三级部门', '') or '').strip()
    if d3 in DEPT_MAP:
        return d3
    return '其他'

def safe_float(v, default=0.0):
    if v is None:
        return default
    try:
        return float(v)
    except:
        return default

def get_days(row):
    """获取欠款天数，优先欠款天数字段，然后计算"""
    days = row.get('欠款天数')
    if days is not None:
        try:
            return int(float(days))
        except:
            pass
    return 0

# ============================================================
# 1. 业绩数据：客户维度汇总（26财年Q1）
# 结构：dept -> seller -> customer -> {perf, collect, debt}
# ============================================================
# 只取已审核状态
perf_customer = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {
    'perf': 0.0, 'collect': 0.0, 'debt': 0.0, 'orders': 0
})))

for row in rows_26:
    if str(row.get('状态', '')).strip() == '作废':
        continue
    dept = get_dept(row)
    seller = str(row.get('销售员名称', '') or '').strip()
    customer = str(row.get('客户名称', '') or '').strip()
    if not customer:
        customer = '未知客户'
    perf_val = safe_float(row.get('业绩总金额')) / 10000
    collect_val = safe_float(row.get('回款金额')) / 10000
    debt_val = safe_float(row.get('欠款金额')) / 10000
    
    perf_customer[dept][seller][customer]['perf'] += perf_val
    perf_customer[dept][seller][customer]['collect'] += collect_val
    perf_customer[dept][seller][customer]['debt'] += debt_val
    perf_customer[dept][seller][customer]['orders'] += 1

# ============================================================
# 2. 欠款数据：客户维度（来自欠款sheet）
# 结构：dept -> seller -> customer -> {debt, d30, d30_90, d90_180, d180, days}
# ============================================================
debt_customer = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {
    'total_debt': 0.0, 'd30': 0.0, 'd30_90': 0.0, 'd90_180': 0.0, 'd180': 0.0,
    'max_days': 0, 'orders': 0
})))

for row in rows_debt:
    if str(row.get('状态', '')).strip() == '作废':
        continue
    dept = get_dept(row)
    seller = str(row.get('销售员名称', '') or '').strip()
    customer = str(row.get('客户名称', '') or '').strip()
    if not customer:
        customer = '未知客户'
    debt_val = safe_float(row.get('欠款金额')) / 10000
    days = get_days(row)
    
    debt_customer[dept][seller][customer]['total_debt'] += debt_val
    debt_customer[dept][seller][customer]['orders'] += 1
    if days > debt_customer[dept][seller][customer]['max_days']:
        debt_customer[dept][seller][customer]['max_days'] = days
    
    # 按天数分类
    if days <= 30:
        debt_customer[dept][seller][customer]['d30'] += debt_val
    elif days <= 90:
        debt_customer[dept][seller][customer]['d30_90'] += debt_val
    elif days <= 180:
        debt_customer[dept][seller][customer]['d90_180'] += debt_val
    else:
        debt_customer[dept][seller][customer]['d180'] += debt_val

# ============================================================
# 3. 回款周期：客户维度
# 从26财年Q1业绩里计算每笔业绩的回款周期（欠款日期到今天 or 回款日期到起始日期）
# ============================================================
TODAY = datetime(2026, 6, 9)
collect_customer = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {
    'total_debt': 0.0, 'total_collect': 0.0,
    'weighted_days': 0.0, 'total_weight': 0.0, 'cycle': 0.0
})))

# 用欠款sheet计算（有业绩日期和欠款天数）
for row in rows_debt:
    if str(row.get('状态', '')).strip() == '作废':
        continue
    dept = get_dept(row)
    seller = str(row.get('销售员名称', '') or '').strip()
    customer = str(row.get('客户名称', '') or '').strip()
    if not customer:
        customer = '未知客户'
    debt_val = safe_float(row.get('欠款金额')) / 10000
    days = get_days(row)
    
    if debt_val > 0 and days > 0:
        collect_customer[dept][seller][customer]['weighted_days'] += debt_val * days
        collect_customer[dept][seller][customer]['total_weight'] += debt_val

# 从业绩表计算回款
for row in rows_26:
    if str(row.get('状态', '')).strip() == '作废':
        continue
    dept = get_dept(row)
    seller = str(row.get('销售员名称', '') or '').strip()
    customer = str(row.get('客户名称', '') or '').strip()
    if not customer:
        customer = '未知客户'
    collect_val = safe_float(row.get('回款金额')) / 10000
    collect_customer[dept][seller][customer]['total_collect'] += collect_val

# 计算加权回款周期
for dept, sellers in collect_customer.items():
    for seller, customers in sellers.items():
        for cust, data in customers.items():
            if data['total_weight'] > 0:
                data['cycle'] = round(data['weighted_days'] / data['total_weight'], 1)
            else:
                data['cycle'] = 0.0

# ============================================================
# 4. 汇总生成JSON结构
# 三级结构：dept -> seller -> [customers]
# ============================================================
def build_customer_detail():
    result = {}
    # 获取所有部门
    all_depts = set(list(perf_customer.keys()) + list(debt_customer.keys()))
    
    for dept in all_depts:
        result[dept] = {}
        # 获取该部门所有销售员
        all_sellers = set(list(perf_customer[dept].keys()) + list(debt_customer[dept].keys()))
        
        for seller in all_sellers:
            # 获取该销售员所有客户（合并业绩和欠款中的客户）
            all_customers = set(
                list(perf_customer[dept][seller].keys()) + 
                list(debt_customer[dept][seller].keys())
            )
            
            customers_list = []
            for cust in all_customers:
                p = perf_customer[dept][seller].get(cust, {})
                d = debt_customer[dept][seller].get(cust, {})
                c = collect_customer[dept][seller].get(cust, {})
                
                perf = round(p.get('perf', 0.0), 2)
                collect = round(p.get('collect', 0.0), 2)
                total_debt = round(d.get('total_debt', 0.0), 2)
                d30 = round(d.get('d30', 0.0), 2)
                d30_90 = round(d.get('d30_90', 0.0), 2)
                d90_180 = round(d.get('d90_180', 0.0), 2)
                d180 = round(d.get('d180', 0.0), 2)
                max_days = d.get('max_days', 0)
                cycle = c.get('cycle', 0.0)
                orders = p.get('orders', d.get('orders', 0))
                
                customers_list.append({
                    'customer': cust,
                    'perf': perf,
                    'collect': collect,
                    'total_debt': total_debt,
                    'd30': d30,
                    'd30_90': d30_90,
                    'd90_180': d90_180,
                    'd180': d180,
                    'max_days': max_days,
                    'cycle': cycle,
                    'orders': orders
                })
            
            # 按业绩额降序排序
            customers_list.sort(key=lambda x: x['perf'], reverse=True)
            result[dept][seller] = customers_list
    
    return result

customer_detail = build_customer_detail()

# 统计
total_customers = sum(
    len(custs) 
    for dept in customer_detail.values() 
    for custs in dept.values()
)
print(f"客户维度数据提取完成:")
print(f"  部门数: {len(customer_detail)}")
print(f"  销售员数: {sum(len(v) for v in customer_detail.values())}")
print(f"  客户总数: {total_customers}")

# 打印示例
for dept, sellers in list(customer_detail.items())[:1]:
    print(f"\n[{dept}] 示例:")
    for seller, custs in list(sellers.items())[:1]:
        print(f"  [{seller}] {len(custs)}个客户:")
        for c in custs[:3]:
            print(f"    {c}")

# 保存
with open(os.path.join(BASE_DIR, 'customer_detail.json'), 'w', encoding='utf-8') as f:
    json.dump(customer_detail, f, ensure_ascii=False, indent=2)

print("\n客户维度数据已保存到 customer_detail.json")
