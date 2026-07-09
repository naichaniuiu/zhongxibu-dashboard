# -*- coding: utf-8 -*-
"""
中西部大区 26财年Q2 数据看板生成器（单文件版）
合并 process_data_v3 + extract_customers_v3 + gen_modal_dashboard 逻辑
数据源: D:/业绩 欠款看板 Q2.xlsx
预估业绩: 6790万元（用户指定，部门按实际比例缩放）
"""
import openpyxl
import json
import os
from datetime import datetime, timedelta
from collections import defaultdict, Counter

# ============================================================
# 配置
# ============================================================
TODAY = datetime(2026, 7, 9)        # 数据基准日
DATA_FILE = 'D:/业绩 欠款看板 Q2.xlsx'
ESTIMATED_TOTAL = 6790.0             # 用户指定的Q2预估业绩（万元）

Q2_START = datetime(2026, 7, 1)
Q2_END   = datetime(2026, 9, 30)
Q2_START_25 = datetime(2025, 7, 1)
Q2_END_25   = datetime(2025, 9, 30)

BLACKLIST = {'支振岗', '李国栋', '白雨'}
KEEP_LIST = {'张宸睿'}

DEPT_MAP = {
    '中西部大区': '中西部大区',
    '中西大区': '中西部大区',
    '华中大区': '中西部大区',
    '华中大区（已封存）': '中西部大区',
    '西南大区': '中西部大区',
    '西南大区（已封存）': '中西部大区',
}

# ============================================================
# 工具函数
# ============================================================
def parse_date(val):
    if not val:
        return None
    s = str(val).strip().replace('\t', '').split(' ')[0]
    if not s or s == 'None':
        return None
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None

def to_wan(val):
    if val is None:
        return 0.0
    try:
        return float(val) / 10000.0
    except (TypeError, ValueError):
        return 0.0

def weighted_avg(items):
    if not items:
        return 0.0
    total_w = sum(x['amount'] for x in items)
    if total_w <= 0:
        return 0.0
    return sum(x['days'] * x['amount'] for x in items) / total_w

def cycle_weighted_avg(items):
    """回款周期 = Σ(每笔认款金额 × 账龄) ÷ 回款总金额
    认款金额: 每笔认款的协同金额
    账龄: 回款日期 - 业绩日期
    分母: 回款总金额（认款协同金额之和）
    """
    if not items:
        return 0.0
    total_payment = sum(x['amount'] for x in items)
    if total_payment <= 0:
        return 0.0
    return sum(x['amount'] * x['days'] for x in items) / total_payment

def load_rows(path, sheet_idx=0):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[sheet_idx]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield dict(zip(headers, row))
    wb.close()

# ============================================================
# 部门映射
# ============================================================
INTERNET_SELLERS = set()

def scan_internet_sellers(path, sheet_idx=0):
    for r in load_rows(path, sheet_idx=sheet_idx):
        dept1 = str(r.get('一级部门') or '').strip().replace('\t', '')
        if dept1 not in DEPT_MAP:
            continue
        dept2 = str(r.get('二级部门') or '').strip().replace('\t', '')
        sub_dept = str(r.get('三级部门') or '').strip().replace('\t', '')
        if '通讯互联网' in dept2 or '通讯互联网' in sub_dept:
            seller = str(r.get('销售员名称') or '').strip().replace('\t', '')
            if seller:
                INTERNET_SELLERS.add(seller)

def normalize_dept2(dept2, sub_dept, seller_name):
    dept2 = str(dept2 or '其他').strip().replace('\t', '')
    sub_dept = str(sub_dept or '').strip().replace('\t', '')
    seller = str(seller_name or '').strip().replace('\t', '')
    if dept2 == '武汉仓' or sub_dept == '武汉仓':
        return None
    if seller in INTERNET_SELLERS:
        return '湖北营销区'
    if '通讯互联网' in dept2 or '通讯互联网' in sub_dept:
        return '湖北营销区'
    if dept2 in ('湖北营销区', '综合管理办公室'):
        return dept2
    if dept2 in ('武汉金融', '武汉能源交通', '武汉基建制造', '武汉通讯互联网') or \
       sub_dept in ('武汉金融行业组', '武汉能源交通行业组', '武汉基建制造行业组', '武汉通讯互联网行业组'):
        return '湖北营销区'
    if dept2 in ('中西销售助理部', '华中用户拓展部'):
        return '湖北营销区'
    if dept2 in ('四川营销区', '重庆营销区'):
        return '综合管理办公室'
    if dept2 in ('成都', '重庆', '郑州', '长沙', '西安') or \
       sub_dept in ('成都站', '重庆站', '郑州站', '长沙站', '西安站'):
        return '综合管理办公室'
    if dept2 == '解决方案部':
        return '湖北营销区'
    return '湖北营销区'

def normalize_sub_dept(dept2, sub_dept, raw_dept2='', seller_name=''):
    dept2 = str(dept2 or '').strip().replace('\t', '')
    sub_dept = str(sub_dept or '其他').strip().replace('\t', '')
    raw_dept2 = str(raw_dept2 or '').strip().replace('\t', '')
    seller = str(seller_name or '').strip().replace('\t', '')
    if seller in ('吴晗', '李国栋'):
        return '其他'
    if raw_dept2 in ('中西销售助理部', '华中用户拓展部', '解决方案部'):
        return '其他'
    if '成都' in sub_dept:
        return '成都站'
    if '重庆' in sub_dept:
        return '重庆站'
    return sub_dept

# ============================================================
# 0. 预扫描
# ============================================================
print('Scanning internet sales sellers...')
scan_internet_sellers(DATA_FILE, sheet_idx=1)
scan_internet_sellers(DATA_FILE, sheet_idx=2)
print(f'  Internet sellers: {len(INTERNET_SELLERS)}')

# ============================================================
# 1. 读取26Q2业绩数据
# ============================================================
print('Loading 26Q2 performance data...')
perf_records = []
for r in load_rows(DATA_FILE, sheet_idx=1):
    dept1 = str(r.get('一级部门') or '').strip().replace('\t', '')
    if dept1 not in DEPT_MAP:
        continue
    d = parse_date(r.get('业绩日期'))
    if not d or not (Q2_START <= d <= Q2_END):
        continue
    # 新版 Sheet1 已包含"是否核算B端业绩"列，按口径过滤
    if str(r.get('是否核算B端业绩') or '').strip() != '是':
        continue
    raw_dept2 = str(r.get('二级部门') or '').strip().replace('\t', '')
    raw_sub_dept = str(r.get('三级部门') or '').strip().replace('\t', '')
    seller_name = str(r.get('销售员名称') or '').strip().replace('\t', '')
    dept2 = normalize_dept2(raw_dept2, raw_sub_dept, seller_name)
    if dept2 is None:
        continue
    sub_dept = normalize_sub_dept(dept2, raw_sub_dept, raw_dept2, seller_name)
    perf_records.append({
        'date': d,
        'order_no': str(r.get('业绩单号') or '').strip().replace('\t', ''),
        'customer_id': str(r.get('客户编号') or '').strip().replace('\t', ''),
        'customer_name': str(r.get('客户名称') or '').strip().replace('\t', ''),
        'seller_no': str(r.get('销售员工号') or '').strip().replace('\t', ''),
        'seller_name': seller_name,
        'dept': dept2,
        'sub_dept': sub_dept,
        'seller_status': str(r.get('销售员状态') or '').strip().replace('\t', ''),
        'perf': to_wan(r.get('业绩总金额')),
        'collect': to_wan(r.get('回款金额')),
        'debt': to_wan(r.get('欠款金额')),
    })
print(f'  26Q2 performance records: {len(perf_records)}')
actual_total = sum(r['perf'] for r in perf_records)
print(f'  26Q2 actual total (before scaling): {actual_total:.2f} wan')

# 数据截止日 = Q2业绩数据中最大业绩日期（即数据更新的最后一天）
_data_dates = [r['date'] for r in perf_records if r.get('date')]
if _data_dates:
    DATA_END = max(_data_dates)
    print(f'  Data end (max perf date): {DATA_END.strftime("%Y-%m-%d")}')
else:
    DATA_END = Q2_END

# ============================================================
# 2. 读取25Q2业绩数据（同比）
# ============================================================
print('Loading 25Q2 performance data...')
perf_records_25 = []
for r in load_rows(DATA_FILE, sheet_idx=0):
    dept1 = str(r.get('一级部门') or '').strip().replace('\t', '')
    if dept1 not in DEPT_MAP:
        continue
    d = parse_date(r.get('业绩日期'))
    if not d or not (Q2_START_25 <= d <= Q2_END_25):
        continue
    raw_dept2 = str(r.get('二级部门') or '').strip().replace('\t', '')
    raw_sub_dept = str(r.get('三级部门') or '').strip().replace('\t', '')
    seller_name = str(r.get('销售员名称') or '').strip().replace('\t', '')
    dept2 = normalize_dept2(raw_dept2, raw_sub_dept, seller_name)
    if dept2 is None:
        continue
    sub_dept = normalize_sub_dept(dept2, raw_sub_dept, raw_dept2, seller_name)
    perf_records_25.append({
        'dept': dept2,
        'seller_name': seller_name,
        'sub_dept': sub_dept,
        'perf': to_wan(r.get('业绩总金额')),
    })
print(f'  25Q2 performance records: {len(perf_records_25)}')

# ============================================================
# 3. 读取欠款数据（整个文件所有行直接求和，无过滤）
# ============================================================
print('Loading debt data...')
debt_records = []
for r in load_rows(DATA_FILE, sheet_idx=2):
    d = parse_date(r.get('业绩日期'))
    raw_dept2 = str(r.get('二级部门') or '').strip().replace('\t', '')
    raw_sub_dept = str(r.get('三级部门') or '').strip().replace('\t', '')
    seller_name = str(r.get('销售员名称') or '').strip().replace('\t', '')
    dept2 = normalize_dept2(raw_dept2, raw_sub_dept, seller_name)
    if dept2 is None:
        continue
    order_no = str(r.get('业绩单号') or '').strip().replace('\t', '')
    debt_val = to_wan(r.get('欠款金额'))
    days = (TODAY - d).days if d else 0
    if days < 0:
        days = 0
    debt_records.append({
        'date': d,
        'order_no': order_no,
        'customer_id': str(r.get('客户编号') or '').strip().replace('\t', ''),
        'customer_name': str(r.get('客户名称') or '').strip().replace('\t', ''),
        'seller_no': str(r.get('销售员工号') or '').strip().replace('\t', ''),
        'seller_name': seller_name,
        'seller_status': str(r.get('销售员状态') or '').strip().replace('\t', ''),
        'dept': dept2,
        'sub_dept': '__pending__',
        'debt': debt_val,
        'days': days,
    })
print(f'  Debt rows: {len(debt_records)}, total debt: {sum(r["debt"] for r in debt_records):.2f}')

# ============================================================
# 3.5 构建订单金额映射表（用于回款周期计算）
# ============================================================
print('Building order amount map...')
order_amount_map = defaultdict(float)  # 业绩单号 -> 业绩总金额(万)
debt_amount_map = defaultdict(float)   # 业绩单号 -> 欠款金额(万)

# 从业绩表读取 业绩总金额（Sheet0=25Q2, Sheet1=26Q2，全量读取不加过滤）
for _si in [0, 1]:
    for r in load_rows(DATA_FILE, sheet_idx=_si):
        on = str(r.get('业绩单号') or '').strip().replace('\t', '')
        if not on:
            continue
        order_amount_map[on] += to_wan(r.get('业绩总金额'))

# 从欠款表读取 欠款金额（Sheet2，全量读取不加过滤）
for r in load_rows(DATA_FILE, sheet_idx=2):
    on = str(r.get('业绩单号') or '').strip().replace('\t', '')
    if not on:
        continue
    debt_amount_map[on] += to_wan(r.get('欠款金额'))

print(f'  Orders from perf sheets: {len(order_amount_map)}')
print(f'  Orders from debt sheet: {len(debt_amount_map)}')

# ============================================================
# 4. 读取认款数据，计算回款周期和回款金额
# ============================================================
# 回款周期公式：回款周期 = Σ(每笔订单金额 × 账龄) ÷ 回款总金额
# 订单金额: 每笔订单的业绩总金额（优先从业绩表获取，fallback=欠款金额+已认款金额）
# 账龄 = 回款日期 - 业绩日期
# 回款总金额 = 认款协同金额之和
print('Loading payment data...')

# 回款周期记录
cycle_records = []
# 回款金额按销售员聚合
seller_collect = defaultdict(float)
for r in load_rows(DATA_FILE, sheet_idx=3):
    if str(r.get('目标认款类型') or '').strip() != '业绩单认款':
        continue
    # 使用认款数据自有的部门信息，不依赖业绩数据映射
    dept1 = str(r.get('一级部门') or '').strip().replace('\t', '')
    if dept1 not in DEPT_MAP:
        continue
    raw_dept2 = str(r.get('二级部门') or '').strip().replace('\t', '')
    raw_sub_dept = str(r.get('三级部门') or '').strip().replace('\t', '')
    seller = str(r.get('销售员名称') or '').strip().replace('\t', '')
    dept2 = normalize_dept2(raw_dept2, raw_sub_dept, seller)
    if dept2 is None:
        continue
    sub_dept = normalize_sub_dept(dept2, raw_sub_dept, raw_dept2, seller)

    perf_date = parse_date(r.get('业绩日期'))
    pay_date = parse_date(r.get('回款日期'))
    if not perf_date or not pay_date:
        continue
    days = (pay_date - perf_date).days
    # 账龄 = 回款日期 - 业绩日期，负数无意义，跳过
    if days < 0:
        continue
    amount = to_wan(r.get('认款协同金额'))
    if amount <= 0:
        continue
    order_no = str(r.get('业绩单号') or '').strip().replace('\t', '')
    cycle_records.append({
        'dept': dept2,
        'sub_dept': sub_dept,
        'seller': seller,
        'days': days,
        'amount': amount,  # 认款协同金额（回款金额）
        'order_no': order_no,
        'order_amount': amount,  # 默认用认款金额，后面会更新为订单金额
        'perf_date': perf_date,
        'pay_date': pay_date,
    })
    seller_collect[seller] += amount

# 计算每笔订单的金额（用于回款周期加权）
# 优先级：1. 业绩表的业绩总金额 > 2. 欠款金额+已认款金额 > 3. 认款协同金额本身
paid_amount_map = defaultdict(float)  # 业绩单号 -> 已认款金额合计(万)
for rec in cycle_records:
    if rec['order_no']:
        paid_amount_map[rec['order_no']] += rec['amount']

matched_count = 0
fallback_count = 0
for rec in cycle_records:
    on = rec['order_no']
    if on and on in order_amount_map and order_amount_map[on] > 0:
        rec['order_amount'] = order_amount_map[on]
        matched_count += 1
    elif on and (debt_amount_map.get(on, 0) + paid_amount_map.get(on, 0)) > 0:
        rec['order_amount'] = debt_amount_map[on] + paid_amount_map[on]
        matched_count += 1
    else:
        rec['order_amount'] = rec['amount']  # fallback: 用认款金额
        fallback_count += 1

avg_cycle = cycle_weighted_avg(cycle_records)
total_collect_amount = sum(r['amount'] for r in cycle_records)
print(f'  Payment records for cycle: {len(cycle_records)}, total amount: {total_collect_amount:.2f} wan, avg cycle: {avg_cycle:.1f}')
print(f'  Order amount matched: {matched_count} ({matched_count/len(cycle_records)*100:.1f}%), fallback: {fallback_count}')

# collect 已从 Sheet1 回款金额直接读取，无需回填

# ============================================================
# 5. 缩放26Q2业绩到用户指定总额 6790万
# ============================================================
print(f'Scaling 26Q2 performance to {ESTIMATED_TOTAL} wan...')
if actual_total > 0:
    scale_factor = ESTIMATED_TOTAL / actual_total
else:
    scale_factor = 1.0
print(f'  Scale factor: {scale_factor:.4f}')

for rec in perf_records:
    rec['perf'] = rec['perf'] * scale_factor

scaled_total = sum(r['perf'] for r in perf_records)
print(f'  Scaled total: {scaled_total:.2f} wan')

# 销售员明细使用未缩放的原始业绩（销售员之间是真实业绩关系，不应按总比例缩放）
perf_records_actual = [{**r, 'perf_orig': r['perf'] / scale_factor if scale_factor > 0 else r['perf']} for r in perf_records]

# ============================================================
# 6. 读取Q2目标
# ============================================================
print('Loading Q2 targets...')
# 从 Sheet4 读取 Q2 目标
target_map = {}  # sub_dept -> target (wan)
wb_tmp = openpyxl.load_workbook(DATA_FILE, data_only=True, read_only=True)
ws4 = wb_tmp[wb_tmp.sheetnames[4]]
for row in ws4.iter_rows(min_row=3, max_row=13, values_only=True):
    if not row[0]:
        continue
    dept_name = str(row[0]).strip()
    q2_target = row[3]  # 26财年Q2 column
    if q2_target and dept_name and dept_name not in ('合计', '中西部大区'):
        try:
            target_map[dept_name] = float(q2_target)
        except (TypeError, ValueError):
            pass
wb_tmp.close()
print(f'  Targets loaded: {target_map}')

# ============================================================
# 7. 销售员维度聚合
# ============================================================
print('Aggregating by seller...')
seller_data = defaultdict(lambda: {
    'perf': 0.0, 'collect': 0.0, 'total_debt': 0.0,
    'd30': 0.0, 'd30_90': 0.0, 'd90_180': 0.0, 'd180': 0.0,
    'dept': '其他', 'sub_dept_list': [],
})

for r in perf_records:
    s = seller_data[r['seller_name']]
    s['perf'] += r['perf']
    s['collect'] += r['collect']
    s['dept'] = r['dept']
    if r.get('sub_dept'):
        s['sub_dept_list'].append(r['sub_dept'])

for r in debt_records:
    s = seller_data[r['seller_name']]
    s['total_debt'] += r['debt']
    s['dept'] = r['dept']
    if r['days'] <= 30:
        s['d30'] += r['debt']
    elif r['days'] <= 90:
        s['d30_90'] += r['debt']
    elif r['days'] <= 180:
        s['d90_180'] += r['debt']
    else:
        s['d180'] += r['debt']

# 按部门聚合销售员
sales_detail_data = defaultdict(list)
sales_cycle_detail = defaultdict(list)
seller_sub_dept = {}
for seller, s in seller_data.items():
    dept = s['dept']
    sub_dept = '其他'
    if s['sub_dept_list']:
        sub_dept = Counter(s['sub_dept_list']).most_common(1)[0][0]
    seller_sub_dept[seller] = sub_dept
    sales_detail_data[dept].append({
        'name': seller,
        'sub_dept': sub_dept,
        'perf': round(s['perf'], 2),
        'collect': round(s['collect'], 2),
        'total_debt': round(s['total_debt'], 2),
        'd30': round(s['d30'], 2),
        'd30_90': round(s['d30_90'], 2),
        'd90_180': round(s['d90_180'], 2),
        'd180': round(s['d180'], 2),
    })
    c_items = [c for c in cycle_records if c['seller'] == seller]
    cycle = cycle_weighted_avg(c_items)
    sales_cycle_detail[dept].append({
        'name': seller,
        'sub_dept': seller_sub_dept.get(seller, '其他'),
        'debt_amt': round(s['total_debt'] * 10000, 2),
        'rec_amt': round(s['collect'] * 10000, 2),
        'cycle': round(cycle, 1),
    })

sales_detail_data = dict(sales_detail_data)
sales_cycle_detail = dict(sales_cycle_detail)

# 回填 debt_records 的 sub_dept
for r in debt_records:
    if r['sub_dept'] == '__pending__':
        r['sub_dept'] = seller_sub_dept.get(r['seller_name'], '其他')

# ============================================================
# 8. 计算 KPI
# ============================================================
print('Calculating KPIs...')
total_perf = sum(r['perf'] for r in perf_records)
total_collect = sum(r['collect'] for r in perf_records)
total_debt = sum(r['debt'] for r in debt_records)

# Q2 目标总额
target_total = sum(target_map.values()) if target_map else ESTIMATED_TOTAL
# 完成率按实际业绩（未缩放）计算
total_completion = round(actual_total / target_total * 100, 1) if target_total > 0 else 0.0

# 在职销售员人数 - 优先使用 Excel 的"在职销售人数" sheet
# 该 sheet 是人力部门提供的官方数据（按三级部门统计 + 大区合计）
# 各部门人数为 26 财年 7 月在职人员状况
active_seller_count = 0
active_seller_count_by_dept = {}
try:
    _hr_wb = openpyxl.load_workbook(DATA_FILE, data_only=True, read_only=True)
    hr_sheet = _hr_wb['在职销售人数']
    for row in hr_sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        dept_name = str(row[0]).strip() if row[0] else ''
        if not dept_name or '合计' in dept_name:
            # 大区合计行：作为总人数
            try:
                v = int(row[1]) if row[1] is not None else 0
                active_seller_count = v
            except (TypeError, ValueError):
                pass
            continue
        # 各部门人数
        try:
            v = int(row[1]) if row[1] is not None else 0
            active_seller_count_by_dept[dept_name] = v
        except (TypeError, ValueError):
            pass
    _hr_wb.close()
    print(f'  HR sheet total: {active_seller_count} | by dept: {active_seller_count_by_dept}')
except Exception as e:
    print(f'  WARN: failed to load HR sheet ({e}), falling back to data-derived count')
    all_active_sellers = set()
    for r in perf_records:
        if r['seller_status'] == '在职' and r['seller_name']:
            all_active_sellers.add(r['seller_name'])
    for r in debt_records:
        if r['seller_status'] == '在职' and r['seller_name']:
            all_active_sellers.add(r['seller_name'])
    active_sellers = (all_active_sellers - BLACKLIST) | KEEP_LIST
    active_seller_count = len(active_sellers)

# 25Q2 总业绩
total_perf_25 = sum(r['perf'] for r in perf_records_25)
total_yoy = round((actual_total - total_perf_25) / total_perf_25 * 100, 1) if total_perf_25 > 0 else None

# 账龄分布
d30 = sum(r['debt'] for r in debt_records if r['days'] <= 30)
d30_90 = sum(r['debt'] for r in debt_records if 30 < r['days'] <= 90)
d90_180 = sum(r['debt'] for r in debt_records if 90 < r['days'] <= 180)
d180 = sum(r['debt'] for r in debt_records if r['days'] > 180)
d_overdue = d30_90 + d90_180 + d180
debt_d90_plus = d90_180 + d180

# 高风险客户
high_risk_customers = []
for r in debt_records:
    if r['debt'] >= 50.0 or r['days'] > 90:
        high_risk_customers.append({
            'customer': r['customer_name'],
            'debt': round(r['debt'], 2),
            'days': r['days'],
            'seller': r['seller_name'],
            'dept': r['dept'],
            'sub_dept': r['sub_dept'],
        })
high_risk_customers.sort(key=lambda x: (-x['debt'], -x['days']))
high_risk_customers = high_risk_customers[:50]

# ============================================================
# 9. 按三级部门聚合
# ============================================================
print('Aggregating by sub-department...')
subdept_perf = defaultdict(lambda: {'perf': 0.0, 'collect': 0.0, 'sales': set(), 'orders': set()})
for r in perf_records:
    key = (r['dept'], r['sub_dept'])
    sd = subdept_perf[key]
    sd['perf'] += r['perf']
    sd['collect'] += r['collect']
    sd['sales'].add(r['seller_name'])
    sd['orders'].add(r['order_no'])

subdept_perf_25 = defaultdict(float)
for r in perf_records_25:
    subdept_perf_25[(r['dept'], r['sub_dept'])] += r['perf']

subdept_debt = defaultdict(lambda: {'d30': 0.0, 'd30_90': 0.0, 'd90_180': 0.0, 'd180': 0.0, 'total_debt': 0.0})
for r in debt_records:
    key = (r['dept'], r['sub_dept'])
    sd = subdept_debt[key]
    sd['total_debt'] += r['debt']
    if r['days'] <= 30:
        sd['d30'] += r['debt']
    elif r['days'] <= 90:
        sd['d30_90'] += r['debt']
    elif r['days'] <= 180:
        sd['d90_180'] += r['debt']
    else:
        sd['d180'] += r['debt']

# 按三级部门聚合回款周期（直接使用认款数据中的部门信息）
subdept_cycle_items = defaultdict(list)
for rec in cycle_records:
    key = (rec['dept'], rec['sub_dept'])
    subdept_cycle_items[key].append(rec)

# 合并三级部门数据
all_subdepts = set(subdept_perf.keys()) | set(subdept_debt.keys()) | set(subdept_perf_25.keys())
subdept_data = []
for key in sorted(all_subdepts):
    dept, sub_dept = key
    p = subdept_perf.get(key, {'perf': 0.0, 'collect': 0.0, 'sales': set(), 'orders': set()})
    d = subdept_debt.get(key, {'d30': 0.0, 'd30_90': 0.0, 'd90_180': 0.0, 'd180': 0.0, 'total_debt': 0.0})
    v25_sub = subdept_perf_25.get(key, 0.0)
    # 实际业绩（未缩放），用于同比和完成率计算
    actual_perf = p['perf'] / scale_factor if scale_factor > 0 else p['perf']
    yoy = round((actual_perf - v25_sub) / v25_sub * 100, 1) if v25_sub > 0 else None
    target = target_map.get(sub_dept, 0.0)
    completion = round(actual_perf / target * 100, 1) if target > 0 else 0.0
    cycle = cycle_weighted_avg(subdept_cycle_items.get(key, []))
    subdept_data.append({
        'dept': dept,
        'sub_dept': sub_dept,
        'key': f'{dept}|{sub_dept}',
        'v26': round(p['perf'], 2),
        'v26_actual': round(actual_perf, 2),
        'v25': round(v25_sub, 2) if v25_sub > 0 else 0,
        'yoy': yoy if yoy is not None else 0,
        'target': round(target, 2),
        'completion': completion,
        'sales': active_seller_count_by_dept.get(sub_dept, len(p['sales'])),  # 优先用HR sheet数据
        'd30': round(d['d30'], 2),
        'd30_90': round(d['d30_90'], 2),
        'd90_180': round(d['d90_180'], 2),
        'd180': round(d['d180'], 2),
        'total_debt': round(d['total_debt'], 2),
        'collect': round(p['collect'], 2),
        'cycle': round(cycle, 1),
    })
subdept_data.sort(key=lambda x: -x['v26'])

# ============================================================
# 10. 客户明细提取
# ============================================================
print('Extracting customer details...')
# dept -> seller -> customer -> dict
cust_result = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {
    'customer': '', 'perf': 0.0, 'collect': 0.0,
    'total_debt': 0.0, 'd30': 0.0, 'd30_90': 0.0, 'd90_180': 0.0, 'd180': 0.0,
    'max_days': 0, 'cycle': 0.0, 'cycle_weight': 0.0, 'cycle_days_weighted': 0.0,
    'orders': 0,
})))

# 业绩数据（销售员明细使用原始未缩放业绩）
for r in perf_records_actual:
    customer = r['customer_name'] or '未知客户'
    c = cust_result[r['dept']][r['seller_name']][customer]
    c['customer'] = customer
    c['sub_dept'] = r['sub_dept']
    c['perf'] += r['perf_orig']
    c['orders'] += 1

# 欠款数据
for r in debt_records:
    customer = r['customer_name'] or '未知客户'
    c = cust_result[r['dept']][r['seller_name']][customer]
    c['customer'] = customer
    c['total_debt'] += r['debt']
    if r['days'] > c['max_days']:
        c['max_days'] = r['days']
    if r['days'] <= 30:
        c['d30'] += r['debt']
    elif r['days'] <= 90:
        c['d30_90'] += r['debt']
    elif r['days'] <= 180:
        c['d90_180'] += r['debt']
    else:
        c['d180'] += r['debt']

# 认款数据
for rec in cycle_records:
    seller = rec['seller']
    # Find customer from payment data - match by order_no
    pass  # Customer-level cycle is approximated from debt+collect

# Build customer_detail.json structure
customer_detail = {}
for dept, sellers in cust_result.items():
    customer_detail[dept] = {}
    for seller, customers in sellers.items():
        customers_list = []
        for customer, data in customers.items():
            # Calculate cycle for customer
            total_weight = data['total_debt'] + data['perf']
            if total_weight > 0 and data['max_days'] > 0:
                # Approximate cycle as max_days * (debt / (debt + perf))
                data['cycle'] = round(data['max_days'] * (data['total_debt'] / total_weight if total_weight > 0 else 0), 1)
            else:
                data['cycle'] = 0.0
            if data['perf'] <= 0 and data['total_debt'] <= 0:
                continue
            customers_list.append({
                'customer': data['customer'],
                'sub_dept': data.get('sub_dept', '其他') or '其他',
                'perf': round(data['perf'], 2),
                'collect': round(data['collect'], 2),
                'total_debt': round(data['total_debt'], 2),
                'd30': round(data['d30'], 2),
                'd30_90': round(data['d30_90'], 2),
                'd90_180': round(data['d90_180'], 2),
                'd180': round(data['d180'], 2),
                'max_days': data['max_days'],
                'cycle': data['cycle'],
                'orders': data['orders'],
            })
        customers_list.sort(key=lambda x: -x['total_debt'])
        if customers_list:
            customer_detail[dept][seller] = customers_list
customer_detail = {k: v for k, v in customer_detail.items() if v}

# ============================================================
# 11. 回款周期极值
# ============================================================
cycles_with_data = [d for d in subdept_data if d['cycle'] > 0]
if cycles_with_data:
    max_cycle_dept = max(cycles_with_data, key=lambda x: x['cycle'])['sub_dept']
    max_cycle = max(d['cycle'] for d in cycles_with_data)
    min_cycle_dept = min(cycles_with_data, key=lambda x: x['cycle'])['sub_dept']
    min_cycle = min(d['cycle'] for d in cycles_with_data)
else:
    max_cycle_dept = ''
    max_cycle = 0.0
    min_cycle_dept = ''
    min_cycle = 0.0
over90_depts = [d['sub_dept'] for d in cycles_with_data if d['cycle'] > 90]

# ============================================================
# 12. 输出 KPI 摘要
# ============================================================
print('\n=== KPI Summary ===')
print(f'  26Q2 Performance: {total_perf:.2f} wan (estimated, actual={actual_total:.2f})')
print(f'  25Q2 Performance: {total_perf_25:.2f} wan')
print(f'  YoY: {total_yoy}%')
print(f'  Q2 Target: {target_total:.1f} wan')
print(f'  Completion: {total_completion}%')
print(f'  Active Sellers: {active_seller_count}')
print(f'  Total Debt: {total_debt:.2f} wan')
print(f'  Overdue (>30d): {d_overdue:.2f} wan')
print(f'  90d+ Debt: {debt_d90_plus:.2f} wan')
print(f'  Avg Cycle: {avg_cycle:.1f} days')
print(f'  Sub-departments: {len(subdept_data)}')

# ============================================================
# 13. 生成 HTML 看板
# ============================================================
print('\nGenerating HTML dashboard...')

# 数据准备
total = {
    'data_date': DATA_END.strftime('%Y-%m-%d'),
    'stat_date': TODAY.strftime('%Y-%m-%d'),
    'gen_date': datetime.now().strftime('%Y-%m-%d'),
    'v26': total_perf,
    'v26_actual': actual_total,
    'v25': total_perf_25,
    'target': target_total,
    'completion': total_completion,
    'yoy': total_yoy if total_yoy is not None else 0,
    'sales': active_seller_count,
    'total_debt': total_debt,
    'overdue': d_overdue,
    'd30': d30,
    'd30_90': d30_90,
    'd90_180': d90_180,
    'd180': d180,
    'collect': total_collect_amount,
    'avg_cycle': avg_cycle,
}

# 按 (dept, sub_dept) 聚合销售员
sales_by_subdept = {}
for dept_name, sellers_map in customer_detail.items():
    sales_by_subdept[dept_name] = {}
    for seller_name, custs in sellers_map.items():
        # 收集该销售员所有客户的 sub_dept，过滤无效值
        valid_sds = [c.get('sub_dept') for c in custs if c.get('sub_dept') and c.get('sub_dept') != '__pending__']
        # 优先使用该销售员在业绩数据中的 sub_dept
        perf_sd = seller_sub_dept.get(seller_name)
        if perf_sd and perf_sd != '其他' and any(sd == perf_sd for sd in valid_sds):
            main_sd = perf_sd
        elif valid_sds:
            sd_counts = {}
            for sd in valid_sds:
                sd_counts[sd] = sd_counts.get(sd, 0) + 1
            main_sd = max(sd_counts, key=sd_counts.get)
        else:
            main_sd = perf_sd or '其他'
        entry = {
            'seller': seller_name,
            'sub_dept': main_sd,
            'perf': round(sum(c.get('perf', 0) for c in custs), 2),
            'collect': round(sum(c.get('collect', 0) for c in custs), 2),
            'total_debt': round(sum(c.get('total_debt', 0) for c in custs), 2),
            'd30': round(sum(c.get('d30', 0) for c in custs), 2),
            'd30_90': round(sum(c.get('d30_90', 0) for c in custs), 2),
            'd90_180': round(sum(c.get('d90_180', 0) for c in custs), 2),
            'd180': round(sum(c.get('d180', 0) for c in custs), 2),
            'cycle': 0,
            'orders': sum(c.get('orders', 0) for c in custs),
            'custs': len(custs),
        }
        sales_by_subdept[dept_name].setdefault(main_sd, []).append(entry)

# 填入 cycle
for dept_name, seller_list in sales_cycle_detail.items():
    if not isinstance(seller_list, list):
        continue
    for s_data in seller_list:
        seller_name = s_data.get('name', '')
        cycle = s_data.get('cycle', 0)
        for sd_key, sellers in sales_by_subdept.get(dept_name, {}).items():
            for s in sellers:
                if s['seller'] == seller_name:
                    s['cycle'] = round(cycle, 1) if cycle else 0

# 客户明细按 sub_dept 聚合
cust_by_subdept = {}
for dept_name, sellers_map in customer_detail.items():
    cust_by_subdept[dept_name] = {}
    for seller_name, custs in sellers_map.items():
        for c in custs:
            sd = c.get('sub_dept', '其他')
            entry = dict(c)
            entry['seller'] = seller_name
            cust_by_subdept[dept_name].setdefault(sd, []).append(entry)

# 部门列表
dept_list = subdept_data

# 高风险客户（90天以上）
risky_customers = []
for dept_name, sellers_map in customer_detail.items():
    for seller_name, custs in sellers_map.items():
        for c in custs:
            d90 = c.get('d90_180', 0) + c.get('d180', 0)
            if d90 > 0:
                risky_customers.append({
                    'customer': c['customer'],
                    'dept': dept_name,
                    'sub_dept': c.get('sub_dept', '其他'),
                    'seller': seller_name,
                    'd90_180': round(c.get('d90_180', 0), 2),
                    'd180': round(c.get('d180', 0), 2),
                    'total_90plus': round(d90, 2),
                    'max_days': c.get('max_days', 0),
                })
risky_customers.sort(key=lambda x: x['total_90plus'], reverse=True)
risky_top15 = risky_customers[:15]

# ========== 渲染辅助函数 ==========
def get_dept_status(completion):
    if completion >= 60:
        return 'badge-good', '较好'
    elif completion >= 30:
        return 'badge-warning', '待关注'
    else:
        return 'badge-down', '严重下滑'

def fmt_yoy(v):
    if v is None or v == 0:
        return '<span class="trend-neutral">-</span>'
    if v > 0:
        return f'<span class="trend-up">&#9650; +{v:.1f}%</span>'
    else:
        return f'<span class="trend-down">&#9660; {v:.1f}%</span>'

def get_debt_status(d):
    total_d = d['total_debt']
    risky = d['d90_180'] + d['d180']
    if risky > 50 or (total_d > 0 and risky/total_d > 0.3):
        return 'badge-down', '高风险'
    elif risky > 10 or total_d > 100:
        return 'badge-warning', '关注'
    else:
        return 'badge-good', '较好'

def get_cycle_status(cycle):
    if cycle > 90:
        return 'badge-down', '需关注', 'negative'
    elif cycle > 60:
        return 'badge-warning', '一般', 'warning'
    else:
        return 'badge-good', '良好', 'highlight'

# 业绩表格行
perf_rows = ''
for d in dept_list:
    badge_cls, badge_text = get_dept_status(d['completion'])
    yoy_html = fmt_yoy(d['yoy'])
    v25_cell = f'{d["v25"]:.2f}' if d['v25'] and d['v25'] > 0 else '-'
    perf_rows += f'''<tr onclick="showSellers('{d["key"]}','perf')" style="cursor:pointer;" title="点击查看销售员业绩明细">
        <td>&#128194; {d["sub_dept"]}</td>
        <td class="highlight">{d["v26_actual"]:.2f}</td>
        <td>{d["target"]:.1f}</td>
        <td>{d["completion"]:.1f}%</td>
        <td>{v25_cell}</td>
        <td>{yoy_html}</td>
        <td>{d["sales"]}</td>
        <td><span class="status-badge {badge_cls}">{badge_text}</span></td>
    </tr>'''

yoy_total_str = f"&#9650; +{total['yoy']:.1f}%" if total['yoy'] > 0 else f"&#9660; {total['yoy']:.1f}%"
v25_total_str = f"{total['v25']:.2f}" if total['v25'] > 0 else '-'

# 欠款表格行
debt_sorted = sorted(dept_list, key=lambda x: x['total_debt'], reverse=True)
debt_rows = ''
for d in debt_sorted:
    badge_cls, badge_text = get_debt_status(d)
    debt_rows += f'''<tr onclick="showSellers('{d["key"]}','debt')" style="cursor:pointer;">
        <td>&#128194; {d["sub_dept"]}</td>
        <td>{d["d30"]:.2f}</td>
        <td>{d["d30_90"]:.2f}</td>
        <td class="warning">{d["d90_180"]:.2f}</td>
        <td class="negative">{d["d180"]:.2f}</td>
        <td class="negative">{d["total_debt"]:.2f}</td>
        <td><span class="status-badge {badge_cls}">{badge_text}</span></td>
    </tr>'''

# 回款周期表格行
cycle_sorted = sorted(dept_list, key=lambda x: x['cycle'], reverse=True)
cycle_rows = ''
for d in cycle_sorted:
    badge_cls, badge_text, val_cls = get_cycle_status(d['cycle'])
    cycle_rows += f'''<tr onclick="showSellers('{d["key"]}','cycle')" style="cursor:pointer;">
        <td>&#128194; {d["sub_dept"]}</td>
        <td>{d["total_debt"]:.2f}</td>
        <td>{d["collect"]:.2f}</td>
        <td class="{val_cls}">{d["cycle"]:.1f}</td>
        <td><span class="status-badge {badge_cls}">{badge_text}</span></td>
    </tr>'''

# 高风险客户行
risky_rows = ''
for i, r in enumerate(risky_top15):
    rank = i + 1
    total_90 = r['total_90plus']
    if total_90 > 50:
        risk_text = '极高风险'
        risk_cls = 'badge-down'
    elif total_90 > 20:
        risk_text = '高风险'
        risk_cls = 'badge-down'
    else:
        risk_text = '关注'
        risk_cls = 'badge-warning'
    risky_rows += f'<tr><td>{rank}</td><td style="text-align:left;">{r["customer"]}（{r["sub_dept"]}，{r["seller"]}）</td><td class="negative">{total_90:.2f}</td><td><span class="status-badge {risk_cls}">{risk_text}</span></td></tr>'

# 超90天部门
over90_list = [d for d in dept_list if d['cycle'] > 90]

# 图表数据
chart_cycle_depts = [d['sub_dept'] for d in cycle_sorted]
chart_cycle_vals = [d['cycle'] for d in cycle_sorted]
chart_cycle_colors = ['#ff4757' if v > 90 else '#ffa502' if v > 60 else '#00ff88' for v in chart_cycle_vals]

perf_depts = [d['sub_dept'] for d in dept_list]
perf_v26 = [d['v26'] for d in dept_list]
perf_v25 = [d['v25'] if d['v25'] and d['v25'] > 0 else 0 for d in dept_list]
perf_pie = [(d['sub_dept'], d['v26']) for d in sorted(dept_list, key=lambda x: x['v26'], reverse=True) if d['v26'] > 0]

debt_pie_labels = ['30天内', '30-90天', '90-180天', '180天以上']
debt_pie_vals = [total['d30'], total['d30_90'], total['d90_180'], total['d180']]
debt_bar_depts = [d['sub_dept'] for d in debt_sorted[:10]]
debt_bar_vals = [d['total_debt'] for d in debt_sorted[:10]]

# JSON 序列化
sales_by_subdept_json = json.dumps(sales_by_subdept, ensure_ascii=False)
cust_by_subdept_json = json.dumps(cust_by_subdept, ensure_ascii=False)
dept_list_json = json.dumps([{
    'dept': d['dept'], 'sub_dept': d['sub_dept'], 'key': d['key'],
    'v26': d['v26'], 'v25': d['v25'], 'yoy': d['yoy'],
    'target': d['target'], 'completion': d['completion'], 'sales': d['sales'],
    'd30': d['d30'], 'd30_90': d['d30_90'], 'd90_180': d['d90_180'], 'd180': d['d180'],
    'total_debt': d['total_debt'], 'collect': d['collect'], 'cycle': d['cycle']
} for d in dept_list], ensure_ascii=False)

# 同比 KPI 卡片
if total['yoy'] > 0:
    yoy_card = f'<div class="value highlight">+{total["yoy"]:.1f}%</div><div class="sub">26Q2实际：{total["v26_actual"]:.2f}万 vs 25Q2：{total["v25"]:.2f}万</div>'
elif total['yoy'] < 0:
    yoy_card = f'<div class="value negative">{total["yoy"]:.1f}%</div><div class="sub">26Q2实际：{total["v26_actual"]:.2f}万 vs 25Q2：{total["v25"]:.2f}万</div>'
else:
    yoy_card = f'<div class="value" style="color:#8892b0;">-</div><div class="sub">26Q2实际：{total["v26_actual"]:.2f}万 vs 25Q2：{total["v25"]:.2f}万</div>'

html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>中西部大区 26财年Q2 数据看板</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.2.0/dist/chartjs-plugin-datalabels.min.js"></script>
    <style>
        * {{ margin:0; padding:0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'PingFang SC', 'Microsoft YaHei', sans-serif; background: #0a0e27; color: #fff; line-height: 1.6; min-height: 100vh; }}
        .container {{ max-width: 1400px; margin: 0 auto; padding: 30px 20px; }}
        .header {{ text-align: center; margin-bottom: 40px; }}
        .header h1 {{ font-size: 2.2em; background: linear-gradient(135deg, #00d4ff 0%, #7b2ff7 100%); -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text; margin-bottom: 10px; }}
        .header p {{ color: #8892b0; font-size: 1.1em; }}

        .kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin-bottom: 30px; }}
        .kpi-card {{ background: linear-gradient(135deg, rgba(255,255,255,0.1) 0%, rgba(255,255,255,0.05) 100%); border-radius: 16px; padding: 24px; text-align: center; border: 1px solid rgba(255,255,255,0.1); transition: transform 0.3s, box-shadow 0.3s; }}
        .kpi-card:hover {{ transform: translateY(-5px); box-shadow: 0 20px 40px rgba(0,212,255,0.15); }}
        .kpi-icon {{ font-size: 2.5em; margin-bottom: 10px; }}
        .kpi-card h3 {{ color: #8892b0; font-size: 0.9em; font-weight: 500; margin-bottom: 8px; }}
        .kpi-card .value {{ font-size: 2.2em; font-weight: 700; color: #00d4ff; }}
        .kpi-card .sub {{ font-size: 0.85em; color: #8892b0; margin-top: 5px; }}

        .highlight {{ color: #00ff88 !important; }}
        .negative {{ color: #ff4757 !important; }}
        .warning {{ color: #ffa502 !important; }}

        .tab-nav {{ display: flex; gap: 10px; margin-bottom: 25px; padding: 8px; background: rgba(255,255,255,0.03); border-radius: 12px; border: 1px solid rgba(255,255,255,0.08); overflow-x: auto; scrollbar-width: none; }}
        .tab-nav::-webkit-scrollbar {{ display: none; }}
        .tab-btn {{ display: flex; align-items: center; gap: 8px; padding: 12px 24px; background: transparent; border: none; border-radius: 8px; color: #8892b0; font-size: 0.95em; font-weight: 500; cursor: pointer; transition: all 0.3s; white-space: nowrap; }}
        .tab-btn:hover {{ background: rgba(255,255,255,0.05); color: #ccd6f6; }}
        .tab-btn.active {{ background: linear-gradient(135deg, #00d4ff 0%, #7b2ff7 100%); color: #fff; box-shadow: 0 4px 15px rgba(0,212,255,0.3); }}

        .tab-content {{ display: none; }}
        .tab-content.active {{ display: block; animation: fadeIn 0.3s ease; }}
        @keyframes fadeIn {{ from {{ opacity: 0; transform: translateY(10px); }} to {{ opacity: 1; transform: translateY(0); }} }}

        .module {{ background: rgba(255,255,255,0.05); border-radius: 20px; padding: 30px; border: 1px solid rgba(255,255,255,0.1); margin-bottom: 25px; }}
        .module-title {{ font-size: 1.3em; margin-bottom: 10px; color: #00d4ff; display: flex; align-items: center; gap: 10px; }}
        .module-desc {{ color: #8892b0; margin-bottom: 20px; font-size: 0.9em; }}

        .charts-section {{ display: grid; grid-template-columns: 1fr 1fr; gap: 25px; margin-bottom: 30px; }}
        .chart-box {{ background: rgba(255,255,255,0.03); border-radius: 16px; padding: 20px; border: 1px solid rgba(255,255,255,0.08); }}
        .chart-box h3 {{ color: #00d4ff; margin-bottom: 15px; font-size: 1.1em; text-align: center; }}
        .chart-container {{ position: relative; height: 320px; }}

        .dept-table {{ width: 100%; border-collapse: collapse; font-size: 0.9em; }}
        .dept-table thead th {{ background: rgba(0,212,255,0.1); color: #00d4ff; padding: 12px 8px; text-align: center; font-weight: 500; border-bottom: 1px solid rgba(0,212,255,0.2); white-space: nowrap; }}
        .dept-table tbody tr {{ border-bottom: 1px solid rgba(255,255,255,0.05); transition: background 0.2s; cursor: pointer; }}
        .dept-table tbody tr:hover {{ background: rgba(255,255,255,0.05); }}
        .dept-table tbody td {{ padding: 11px 8px; text-align: center; color: #ccd6f6; }}
        .dept-table tbody td:first-child {{ text-align: left; padding-left: 12px; }}
        .dept-table tfoot tr {{ background: rgba(0,212,255,0.08); border-top: 2px solid rgba(0,212,255,0.3); }}
        .dept-table tfoot td {{ padding: 12px 8px; text-align: center; color: #00d4ff; font-weight: 600; }}
        .dept-table tfoot td:first-child {{ text-align: left; padding-left: 12px; }}

        .trend-down {{ color: #ff4757; font-weight: 600; }}
        .trend-up {{ color: #00ff88; font-weight: 600; }}
        .trend-neutral {{ color: #ffa502; }}
        .status-badge {{ padding: 3px 10px; border-radius: 12px; font-size: 0.8em; font-weight: 600; }}
        .badge-down {{ background: rgba(255,71,87,0.2); color: #ff4757; border: 1px solid rgba(255,71,87,0.3); }}
        .badge-warning {{ background: rgba(255,165,2,0.2); color: #ffa502; border: 1px solid rgba(255,165,2,0.3); }}
        .badge-new {{ background: rgba(0,255,136,0.2); color: #00ff88; border: 1px solid rgba(0,255,136,0.3); }}
        .badge-good {{ background: rgba(0,212,255,0.2); color: #00d4ff; border: 1px solid rgba(0,212,255,0.3); }}

        .modal-overlay {{ display: none; position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.7); z-index: 1000; }}
        .modal-overlay.active {{ display: flex; align-items: center; justify-content: center; }}
        .modal {{ background: #1a2040; border-radius: 20px; padding: 30px; max-width: 1000px; width: 92%; max-height: 85vh; overflow-y: auto; border: 1px solid rgba(0,212,255,0.3); }}
        .modal-header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px; }}
        .modal-title {{ color: #00d4ff; font-size: 1.2em; font-weight: 600; }}
        .modal-close {{ background: rgba(255,255,255,0.1); border: none; color: #fff; width: 30px; height: 30px; border-radius: 50%; cursor: pointer; font-size: 1.2em; display: flex; align-items: center; justify-content: center; }}
        .modal-close:hover {{ background: rgba(255,71,87,0.3); }}

        @media (max-width: 768px) {{ .charts-section {{ grid-template-columns: 1fr; }} .kpi-grid {{ grid-template-columns: repeat(2, 1fr); }} .dept-table {{ font-size: 0.8em; }} }}
    </style>
</head>
<body>
<div class="container">
    <div class="header">
        <h1>中西部大区 26财年Q2 数据看板</h1>
        <p>数据截止：{total['data_date']} &nbsp;|&nbsp; 统计基日：{total['stat_date']} &nbsp;|&nbsp; 生成于：{total['gen_date']}</p>
    </div>

    <!-- 核心KPI -->
    <div class="kpi-grid">
        <div class="kpi-card">
            <div class="kpi-icon">&#128176;</div>
            <h3>26Q2 预估业绩</h3>
            <div class="value">{total['v26']:.2f}<span style="font-size:0.5em;">万</span></div>
            <div class="sub">目标：{total['target']:.1f}万 | 已完成：{total['v26_actual']:.2f}万</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-icon">&#127919;</div>
            <h3>Q2目标完成率</h3>
            <div class="value {'negative' if total['completion'] < 100 else 'highlight'}">{total['completion']:.1f}%</div>
            <div class="sub">实际已完成 {total['v26_actual']:.2f}万 / 目标 {total['target']:.1f}万</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-icon">&#128200;</div>
            <h3>同比25Q2</h3>
            {yoy_card}
        </div>
        <div class="kpi-card">
            <div class="kpi-icon">&#128101;</div>
            <h3>在职销售员总数</h3>
            <div class="value" style="color:#ffa502;">{total['sales']}</div>
            <div class="sub">人</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-icon">&#9888;&#65039;</div>
            <h3>逾期欠款总额</h3>
            <div class="value negative">{total['overdue']:.2f}<span style="font-size:0.5em;">万</span></div>
            <div class="sub">总欠款：{total['total_debt']:.2f}万 | 90天以上：{total['d90_180']+total['d180']:.2f}万</div>
        </div>
    </div>

    <!-- Tab导航 -->
    <div class="tab-nav">
        <button class="tab-btn active" onclick="switchTab('performance', this)"><span class="tab-icon">&#128202;</span>业绩分析</button>
        <button class="tab-btn" onclick="switchTab('debt', this)"><span class="tab-icon">&#128179;</span>欠款分析</button>
        <button class="tab-btn" onclick="switchTab('collection', this)"><span class="tab-icon">&#9201;&#65039;</span>平均回款周期分析</button>
    </div>

    <!-- ===== 业绩分析 Tab ===== -->
    <div id="tab-performance" class="tab-content active">
        <div class="module">
            <h2 class="module-title">&#128202; 26财年Q2 三级部门业绩总览</h2>
            <div class="charts-section">
                <div class="chart-box">
                    <h3>25Q2 vs 26Q2 三级部门业绩对比（万元）</h3>
                    <div class="chart-container"><canvas id="perfBarChart"></canvas></div>
                </div>
                <div class="chart-box">
                    <h3>26Q2 三级部门业绩占比分布</h3>
                    <div class="chart-container"><canvas id="perfPieChart"></canvas></div>
                </div>
            </div>
            <p class="module-desc" style="color:#00d4ff;font-weight:600;">&#128281; 点击任意三级部门行 → 查看该部门销售员明细</p>
            <table class="dept-table">
                <thead>
                    <tr>
                        <th>三级部门</th>
                        <th>26Q2上线业绩(万)</th>
                        <th>Q2目标(万)</th>
                        <th>实际完成率</th>
                        <th>25Q2同期(万)</th>
                        <th>同比%</th>
                        <th>销售员数</th>
                        <th>状态</th>
                    </tr>
                </thead>
                <tbody>
                    {perf_rows}
                </tbody>
                <tfoot>
                    <tr>
                        <td>合计</td><td>{total['v26_actual']:.2f}</td><td>{total['target']:.1f}</td><td>{total['completion']:.1f}%</td>
                        <td>{v25_total_str}</td><td class="{'trend-up' if total['yoy'] > 0 else 'trend-down'}">{yoy_total_str}</td><td>{total['sales']}</td><td></td>
                    </tr>
                </tfoot>
            </table>
        </div>
    </div>

    <!-- ===== 欠款分析 Tab ===== -->
    <div id="tab-debt" class="tab-content">
        <div class="module">
            <h2 class="module-title">&#128179; 欠款分析总览</h2>

            <div class="kpi-grid" style="margin-bottom:25px;">
                <div class="kpi-card">
                    <div class="kpi-icon">&#128203;</div>
                    <h3>欠款总额</h3>
                    <div class="value negative">{total['total_debt']:.2f}<span style="font-size:0.5em;">万</span></div>
                    <div class="sub">全大区欠款合计</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-icon">&#128994;</div>
                    <h3>30天内</h3>
                    <div class="value highlight">{total['d30']:.2f}<span style="font-size:0.5em;">万</span></div>
                    <div class="sub">占比 {total['d30']/total['total_debt']*100:.1f}%</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-icon">&#128992;</div>
                    <h3>30-90天</h3>
                    <div class="value warning">{total['d30_90']:.2f}<span style="font-size:0.5em;">万</span></div>
                    <div class="sub">占比 {total['d30_90']/total['total_debt']*100:.1f}%</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-icon">&#128308;</div>
                    <h3>90天以上</h3>
                    <div class="value negative">{total['d90_180']+total['d180']:.2f}<span style="font-size:0.5em;">万</span></div>
                    <div class="sub">90-180天：{total['d90_180']:.2f}万 | 180天+：{total['d180']:.2f}万</div>
                </div>
            </div>

            <div class="charts-section">
                <div class="chart-box">
                    <h3>逾期欠款金额占比分布</h3>
                    <div class="chart-container"><canvas id="debtPieChart"></canvas></div>
                </div>
                <div class="chart-box">
                    <h3>各三级部门欠款总额排名（万元）</h3>
                    <div class="chart-container"><canvas id="debtBarChart"></canvas></div>
                </div>
            </div>

            <h3 style="color:#00d4ff;margin-bottom:15px;font-size:1.1em;">各三级部门分账龄欠款明细 <span style="color:#8892b0;font-size:0.8em;font-weight:normal;">（点击部门查看销售员欠款明细）</span></h3>
            <table class="dept-table">
                <thead>
                    <tr>
                        <th>三级部门（点击查看明细）</th>
                        <th>30天内(万)</th>
                        <th>30-90天(万)</th>
                        <th>90-180天(万)</th>
                        <th>180天以上(万)</th>
                        <th>合计(万)</th>
                        <th>状态</th>
                    </tr>
                </thead>
                <tbody>
                    {debt_rows}
                </tbody>
                <tfoot>
                    <tr><td>合计</td><td>{total['d30']:.2f}</td><td>{total['d30_90']:.2f}</td><td>{total['d90_180']:.2f}</td><td>{total['d180']:.2f}</td><td>{total['total_debt']:.2f}</td><td></td></tr>
                </tfoot>
            </table>

            <h3 style="color:#ff4757;margin:25px 0 15px;font-size:1.1em;">&#128680; 高风险客户（90天以上欠款，前15名）</h3>
            <table class="dept-table">
                <thead>
                    <tr><th>排名</th><th>客户名称（三级部门，销售员）</th><th>欠款金额(万)</th><th>风险等级</th></tr>
                </thead>
                <tbody>
                    {risky_rows}
                </tbody>
            </table>
        </div>
    </div>

    <!-- ===== 平均回款周期分析 Tab ===== -->
    <div id="tab-collection" class="tab-content">
        <div class="module">
            <h2 class="module-title">&#9201;&#65039; 平均回款周期分析</h2>
            <div class="kpi-grid" style="margin-bottom:25px;">
                <div class="kpi-card">
                    <div class="kpi-icon">&#128202;</div>
                    <h3>全大区平均回款周期</h3>
                    <div class="value warning">{total['avg_cycle']:.1f}</div>
                    <div class="sub">天（认款金额加权平均）</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-icon">&#128308;</div>
                    <h3>最长三级部门回款周期</h3>
                    <div class="value negative">{max(cycle_sorted[0]['cycle'] if cycle_sorted else 0, 0):.1f}</div>
                    <div class="sub">{cycle_sorted[0]['sub_dept'] if cycle_sorted else '-'}</div>
                </div>
                <div class="kpi-card">
                    <div class="kpi-icon">&#128994;</div>
                    <h3>最短三级部门回款周期</h3>
                    <div class="value highlight">{cycle_sorted[-1]['cycle'] if cycle_sorted else 0:.1f}</div>
                    <div class="sub">{cycle_sorted[-1]['sub_dept'] if cycle_sorted else '-'}</div>
                </div>
                <div class="kpi-card" onclick="showOver90Depts()" style="cursor:pointer;" title="点击查看回款周期大于90天的三级部门">
                    <div class="kpi-icon">&#9888;&#65039;</div>
                    <h3>超90天三级部门数</h3>
                    <div class="value negative">{len(over90_list)}</div>
                    <div class="sub">回款周期大于90天的部门（点击查看）</div>
                </div>
            </div>

            <div style="background:rgba(255, 255, 255, 0.03);border-radius:12px;padding:15px 20px;margin-bottom:20px;border:1px solid rgba(255, 255, 255, 0.08);">
                <div style="display:flex;align-items:center;gap:8px;margin-bottom:10px;">
                    <span style="font-weight:600;color:#00d4ff;">三级部门平均回款周期（天）</span>
                    <span style="color:#8892b0;font-size:0.9em;">| 回款周期计算 = &Sigma;(每笔认款金额 &times; 账龄) &divide; 回款总金额 &nbsp; 账龄 = 回款日期 - 业绩日期</span>
                </div>
                <div style="display:flex;gap:20px;flex-wrap:wrap;">
                    <div style="display:flex;align-items:center;gap:6px;"><span style="width:10px;height:10px;border-radius:50%;background:#00ff88;display:inline-block;"></span><span style="color:#ccd6f6;font-size:0.9em;">&le;60天（良好）</span></div>
                    <div style="display:flex;align-items:center;gap:6px;"><span style="width:10px;height:10px;border-radius:50%;background:#ffa502;display:inline-block;"></span><span style="color:#ccd6f6;font-size:0.9em;">61-90天（一般）</span></div>
                    <div style="display:flex;align-items:center;gap:6px;"><span style="width:10px;height:10px;border-radius:50%;background:#ff4757;display:inline-block;"></span><span style="color:#ccd6f6;font-size:0.9em;">&gt;90天（需关注）</span></div>
                </div>
            </div>

            <div class="chart-box" style="margin-bottom:25px;">
                <h3>&#128202; 三级部门平均回款周期排名</h3>
                <div class="chart-container" style="height:380px;"><canvas id="cycleChart"></canvas></div>
            </div>

            <table class="dept-table">
                <thead>
                    <tr>
                        <th>三级部门（点击查看明细）</th>
                        <th>欠款总额(万)</th>
                        <th>回款金额(万)</th>
                        <th>回款周期(天)</th>
                        <th>状态</th>
                    </tr>
                </thead>
                <tbody>
                    {cycle_rows}
                </tbody>
                <tfoot>
                    <tr><td>合计</td><td>{total['total_debt']:.2f}</td><td>{total['collect']:.2f}</td><td>{total['avg_cycle']:.1f}</td><td></td></tr>
                </tfoot>
            </table>
        </div>
    </div>
</div>

<!-- 销售员/客户明细弹窗 -->
<div class="modal-overlay" id="drillModal" onclick="if(event.target===this)closeDrillModal()" style="z-index:1001;">
    <div style="background:#1a2040;border-radius:20px;padding:30px;max-width:1100px;width:95%;max-height:85vh;overflow-y:auto;border:1px solid rgba(0,212,255,0.3);">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;">
            <div class="modal-title" id="drillModalTitle">明细</div>
            <button class="modal-close" onclick="closeDrillModal()">&#10005;</button>
        </div>
        <div id="drillTableContainer"></div>
    </div>
</div>

<script>
const deptData = {dept_list_json};
const salesBySubDept = {sales_by_subdept_json};
const custBySubDept = {cust_by_subdept_json};

let currentTabType = 'perf';
function switchTab(id, btn) {{
    document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    document.getElementById('tab-' + id).classList.add('active');
    btn.classList.add('active');
    if (id === 'performance') currentTabType = 'perf';
    else if (id === 'debt') currentTabType = 'debt';
    else if (id === 'collection') currentTabType = 'cycle';
}}

function parseKey(key) {{
    const idx = key.indexOf('|');
    return {{ dept: key.substring(0, idx), subDept: key.substring(idx + 1) }};
}}

function showSellers(key, type) {{
    const {{ dept, subDept }} = parseKey(key);
    if (type === 'perf') renderSellerPerf(dept, subDept);
    else if (type === 'debt') renderSellerDebt(dept, subDept);
    else if (type === 'cycle') renderSellerCycle(dept, subDept);
    else renderSellerPerf(dept, subDept);
}}

function renderSellerPerf(dept, subDept) {{
    const sellers = salesBySubDept[dept] && salesBySubDept[dept][subDept] ? [...salesBySubDept[dept][subDept]] : [];
    sellers.sort((a, b) => b.perf - a.perf);
    const rows = sellers.map(s => {{
        const color = s.perf <= 0 ? '#8892b0' : s.perf < 50 ? '#ffa502' : '#00ff88';
        const status = s.perf <= 0 ? '&#11030; 无业绩' : s.perf < 50 ? '&#128992; 较低' : '&#128994; 正常';
        return `<tr style="border-bottom:1px solid rgba(255,255,255,0.05);cursor:pointer;" onclick="showSellerCustomers('${{dept}}','${{subDept}}','${{s.seller}}','perf')" title="点击查看客户明细">
            <td style="padding:8px 6px;color:#ccd6f6;text-align:left;">&#128203; ${{s.seller}}</td>
            <td style="padding:8px 6px;color:${{color}};text-align:right;font-weight:600;">${{(s.perf || 0).toFixed(2)}}</td>
            <td style="padding:8px 6px;color:#00ff88;text-align:right;">${{(s.collect || 0).toFixed(2)}}</td>
            <td style="padding:8px 6px;color:#8892b0;text-align:center;">${{s.orders || 0}}</td>
            <td style="padding:8px 6px;color:#ccd6f6;text-align:center;">${{status}}</td>
        </tr>`;
    }}).join('');
    const tp = sellers.reduce((s, x) => s + (x.perf || 0), 0).toFixed(2);
    const tc = sellers.reduce((s, x) => s + (x.collect || 0), 0).toFixed(2);
    document.getElementById('drillModalTitle').innerHTML = `&#128202; ${{subDept}} - 销售员业绩明细 <span style="font-size:0.7em;color:#8892b0;">（共${{sellers.length}}人）</span>`;
    document.getElementById('drillTableContainer').innerHTML = `<table style="width:100%;border-collapse:collapse;font-size:0.82em;">
        <thead><tr style="background:rgba(0,212,255,0.12);">
            <th style="padding:8px 6px;color:#00d4ff;text-align:left;">销售员</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:right;">26Q2业绩(万)</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:right;">回款(万)</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:center;">订单数</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:center;">状态</th>
        </tr></thead>
        <tbody>${{rows}}
        <tr style="background:rgba(0,212,255,0.08);font-weight:600;">
            <td style="padding:8px 6px;color:#00d4ff;">合计（${{sellers.length}}人）</td>
            <td style="padding:8px 6px;color:#00ff88;text-align:right;">${{tp}}</td>
            <td style="padding:8px 6px;color:#00ff88;text-align:right;">${{tc}}</td>
            <td colspan="2"></td>
        </tr></tbody></table>
        <p style="color:#8892b0;font-size:0.82em;margin-top:8px;">点击销售员行可查看其客户明细</p>`;
    document.getElementById('drillModal').classList.add('active');
}}

function renderSellerDebt(dept, subDept) {{
    const sellers = salesBySubDept[dept] && salesBySubDept[dept][subDept] ? [...salesBySubDept[dept][subDept]] : [];
    sellers.sort((a, b) => b.total_debt - a.total_debt);
    const rows = sellers.map(s => {{
        const dc = s.total_debt > 50 ? '#ff4757' : s.total_debt > 20 ? '#ffa502' : '#00ff88';
        const risky = (s.d90_180 || 0) + (s.d180 || 0);
        const status = risky > 20 ? '&#128308; 高风险' : risky > 5 ? '&#128992; 关注' : '&#128994; 较好';
        return `<tr style="border-bottom:1px solid rgba(255,255,255,0.05);cursor:pointer;" onclick="showSellerCustomers('${{dept}}','${{subDept}}','${{s.seller}}','debt')" title="点击查看客户欠款明细">
            <td style="padding:8px 6px;color:#ccd6f6;text-align:left;">&#128203; ${{s.seller}}</td>
            <td style="padding:8px 6px;color:${{dc}};text-align:right;font-weight:600;">${{(s.total_debt || 0).toFixed(2)}}</td>
            <td style="padding:8px 6px;color:#00ff88;text-align:right;">${{(s.d30 || 0).toFixed(2)}}</td>
            <td style="padding:8px 6px;color:#ffa502;text-align:right;">${{(s.d30_90 || 0).toFixed(2)}}</td>
            <td style="padding:8px 6px;color:${{(s.d90_180||0)>0?'#ff4757':'#8892b0'}};text-align:right;">${{(s.d90_180 || 0).toFixed(2)}}</td>
            <td style="padding:8px 6px;color:${{(s.d180||0)>0?'#ff4757':'#8892b0'}};text-align:right;">${{(s.d180 || 0).toFixed(2)}}</td>
            <td style="padding:8px 6px;color:#ccd6f6;text-align:center;">${{status}}</td>
        </tr>`;
    }}).join('');
    const td = sellers.reduce((s, x) => s + (x.total_debt || 0), 0).toFixed(2);
    document.getElementById('drillModalTitle').innerHTML = `&#128179; ${{subDept}} - 销售员欠款明细 <span style="font-size:0.7em;color:#8892b0;">（共${{sellers.length}}人）</span>`;
    document.getElementById('drillTableContainer').innerHTML = `<table style="width:100%;border-collapse:collapse;font-size:0.8em;">
        <thead><tr style="background:rgba(0,212,255,0.12);">
            <th style="padding:8px 6px;color:#00d4ff;text-align:left;">销售员</th>
            <th style="padding:8px 6px;color:#ff4757;text-align:right;">合计欠款(万)</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:right;">30天内</th>
            <th style="padding:8px 6px;color:#ffa502;text-align:right;">30-90天</th>
            <th style="padding:8px 6px;color:#ff4757;text-align:right;">90-180天</th>
            <th style="padding:8px 6px;color:#ff4757;text-align:right;">180天以上</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:center;">状态</th>
        </tr></thead>
        <tbody>${{rows}}
        <tr style="background:rgba(0,212,255,0.08);font-weight:600;">
            <td style="padding:8px 6px;color:#00d4ff;">合计（${{sellers.length}}人）</td>
            <td style="padding:8px 6px;color:#ff4757;text-align:right;">${{td}}</td>
            <td colspan="5"></td>
        </tr></tbody></table>
        <p style="color:#8892b0;font-size:0.82em;margin-top:8px;">点击销售员行可查看其客户欠款明细</p>`;
    document.getElementById('drillModal').classList.add('active');
}}

function renderSellerCycle(dept, subDept) {{
    const sellers = salesBySubDept[dept] && salesBySubDept[dept][subDept] ? [...salesBySubDept[dept][subDept]] : [];
    sellers.sort((a, b) => (b.cycle || 0) - (a.cycle || 0));
    const rows = sellers.map(s => {{
        const cycle = s.cycle || 0;
        const cycleStr = cycle > 0 ? cycle.toFixed(1) : '-';
        const cc = cycle <= 0 ? '#8892b0' : cycle > 90 ? '#ff4757' : cycle > 60 ? '#ffa502' : '#00ff88';
        const status = cycle <= 0 ? '&#11030; 无数据' : cycle > 90 ? '&#128308; 需关注' : cycle > 60 ? '&#128992; 偏高' : '&#128994; 正常';
        return `<tr style="border-bottom:1px solid rgba(255,255,255,0.05);cursor:pointer;" onclick="showSellerCustomers('${{dept}}','${{subDept}}','${{s.seller}}','cycle')" title="点击查看客户回款周期明细">
            <td style="padding:8px 6px;color:#ccd6f6;text-align:left;">&#128203; ${{s.seller}}</td>
            <td style="padding:8px 6px;color:#00ff88;text-align:right;">${{(s.collect || 0).toFixed(2)}}</td>
            <td style="padding:8px 6px;color:#ffa502;text-align:right;">${{(s.total_debt || 0).toFixed(2)}}</td>
            <td style="padding:8px 6px;color:${{cc}};text-align:right;font-weight:600;">${{cycleStr}}</td>
            <td style="padding:8px 6px;color:#ccd6f6;text-align:center;">${{status}}</td>
        </tr>`;
    }}).join('');
    document.getElementById('drillModalTitle').innerHTML = `&#9201;&#65039; ${{subDept}} - 销售员回款周期明细 <span style="font-size:0.7em;color:#8892b0;">（共${{sellers.length}}人）</span>`;
    document.getElementById('drillTableContainer').innerHTML = `<table style="width:100%;border-collapse:collapse;font-size:0.82em;">
        <thead><tr style="background:rgba(0,212,255,0.12);">
            <th style="padding:8px 6px;color:#00d4ff;text-align:left;">销售员</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:right;">回款(万)</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:right;">欠款(万)</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:right;">回款周期(天)</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:center;">状态</th>
        </tr></thead>
        <tbody>${{rows}}</tbody></table>
        <p style="color:#8892b0;font-size:0.82em;margin-top:8px;">点击销售员行可查看其客户回款周期明细</p>`;
    document.getElementById('drillModal').classList.add('active');
}}

function showSellerCustomers(dept, subDept, seller, type) {{
    if (type === 'perf') renderSellerCustPerf(dept, subDept, seller);
    else if (type === 'debt') renderSellerCustDebt(dept, subDept, seller);
    else if (type === 'cycle') renderSellerCustCycle(dept, subDept, seller);
    else renderSellerCustPerf(dept, subDept, seller);
}}

function renderSellerCustPerf(dept, subDept, seller) {{
    let custList = [];
    if (custBySubDept[dept] && custBySubDept[dept][subDept]) {{
        custList = custBySubDept[dept][subDept].filter(c => c.seller === seller);
    }}
    custList.sort((a, b) => b.perf - a.perf);
    const rows = custList.map(c => {{
        const color = c.perf <= 0 ? '#8892b0' : c.perf < 2 ? '#ffa502' : '#00ff88';
        const status = c.perf <= 0 ? '&#11030; 无业绩' : c.perf < 5 ? '&#128992; 较低' : '&#128994; 正常';
        return `<tr style="border-bottom:1px solid rgba(255,255,255,0.05);">
            <td style="padding:7px 5px;color:#ccd6f6;max-width:300px;word-break:break-all;">${{c.customer}}</td>
            <td style="padding:7px 5px;color:${{color}};text-align:right;font-weight:600;">${{(c.perf || 0).toFixed(2)}}</td>
            <td style="padding:7px 5px;color:#00ff88;text-align:right;">${{(c.collect || 0).toFixed(2)}}</td>
            <td style="padding:7px 5px;color:#8892b0;text-align:center;">${{c.orders || 0}}</td>
            <td style="padding:7px 5px;color:#ccd6f6;text-align:center;">${{status}}</td>
        </tr>`;
    }}).join('');
    const tp = custList.reduce((s, c) => s + (c.perf || 0), 0).toFixed(2);
    const tc = custList.reduce((s, c) => s + (c.collect || 0), 0).toFixed(2);
    const backBtn = `<button onclick="renderSellerPerf('${{dept}}','${{subDept}}')" style="background:rgba(255,255,255,0.1);border:1px solid rgba(255,255,255,0.2);color:#ccd6f6;padding:6px 14px;border-radius:8px;cursor:pointer;font-size:0.85em;margin-bottom:12px;">&larr; 返回销售员列表</button>`;
    document.getElementById('drillModalTitle').innerHTML = `&#128202; ${{seller}} - 客户业绩明细 <span style="font-size:0.7em;color:#8892b0;">（${{subDept}}）</span>`;
    document.getElementById('drillTableContainer').innerHTML = backBtn + `<table style="width:100%;border-collapse:collapse;font-size:0.82em;">
        <thead><tr style="background:rgba(0,212,255,0.12);">
            <th style="padding:8px 6px;color:#00d4ff;text-align:left;">客户名称</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:right;">26Q2业绩(万)</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:right;">回款(万)</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:center;">订单数</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:center;">状态</th>
        </tr></thead>
        <tbody>${{rows}}
        <tr style="background:rgba(0,212,255,0.08);font-weight:600;">
            <td style="padding:8px 6px;color:#00d4ff;">合计（${{custList.length}}个客户）</td>
            <td style="padding:8px 6px;color:#00ff88;text-align:right;">${{tp}}</td>
            <td style="padding:8px 6px;color:#00ff88;text-align:right;">${{tc}}</td>
            <td colspan="2"></td>
        </tr></tbody></table>`;
    document.getElementById('drillModal').classList.add('active');
}}

function renderSellerCustDebt(dept, subDept, seller) {{
    let custList = [];
    if (custBySubDept[dept] && custBySubDept[dept][subDept]) {{
        custList = custBySubDept[dept][subDept].filter(c => c.seller === seller);
    }}
    custList.sort((a, b) => b.total_debt - a.total_debt);
    const rows = custList.map(c => {{
        const dc = c.total_debt > 50 ? '#ff4757' : c.total_debt > 20 ? '#ffa502' : '#00ff88';
        const risky = (c.d90_180 || 0) + (c.d180 || 0);
        const status = risky > 20 ? '&#128308; 高风险' : risky > 5 ? '&#128992; 关注' : '&#128994; 较好';
        const daysText = c.max_days > 0 ? `最长${{c.max_days}}天` : '-';
        return `<tr style="border-bottom:1px solid rgba(255,255,255,0.05);">
            <td style="padding:7px 5px;color:#ccd6f6;max-width:280px;word-break:break-all;">${{c.customer}}</td>
            <td style="padding:7px 5px;color:${{dc}};text-align:right;font-weight:600;">${{(c.total_debt || 0).toFixed(2)}}</td>
            <td style="padding:7px 5px;color:#00ff88;text-align:right;">${{(c.d30 || 0).toFixed(2)}}</td>
            <td style="padding:7px 5px;color:#ffa502;text-align:right;">${{(c.d30_90 || 0).toFixed(2)}}</td>
            <td style="padding:7px 5px;color:${{(c.d90_180||0)>0?'#ff4757':'#8892b0'}};text-align:right;">${{(c.d90_180 || 0).toFixed(2)}}</td>
            <td style="padding:7px 5px;color:${{(c.d180||0)>0?'#ff4757':'#8892b0'}};text-align:right;">${{(c.d180 || 0).toFixed(2)}}</td>
            <td style="padding:7px 5px;color:#8892b0;text-align:center;font-size:0.9em;">${{daysText}}</td>
            <td style="padding:7px 5px;color:#ccd6f6;text-align:center;">${{status}}</td>
        </tr>`;
    }}).join('');
    const td = custList.reduce((s, c) => s + (c.total_debt || 0), 0).toFixed(2);
    const backBtn = `<button onclick="renderSellerDebt('${{dept}}','${{subDept}}')" style="background:rgba(255,255,255,0.1);border:1px solid rgba(255,255,255,0.2);color:#ccd6f6;padding:6px 14px;border-radius:8px;cursor:pointer;font-size:0.85em;margin-bottom:12px;">&larr; 返回销售员列表</button>`;
    document.getElementById('drillModalTitle').innerHTML = `&#128179; ${{seller}} - 客户欠款明细 <span style="font-size:0.7em;color:#8892b0;">（${{subDept}}）</span>`;
    document.getElementById('drillTableContainer').innerHTML = backBtn + `<table style="width:100%;border-collapse:collapse;font-size:0.78em;">
        <thead><tr style="background:rgba(0,212,255,0.12);">
            <th style="padding:8px 6px;color:#00d4ff;text-align:left;">客户名称</th>
            <th style="padding:8px 6px;color:#ff4757;text-align:right;">合计欠款(万)</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:right;">30天内</th>
            <th style="padding:8px 6px;color:#ffa502;text-align:right;">30-90天</th>
            <th style="padding:8px 6px;color:#ff4757;text-align:right;">90-180天</th>
            <th style="padding:8px 6px;color:#ff4757;text-align:right;">180天以上</th>
            <th style="padding:8px 6px;color:#8892b0;text-align:center;">最长天数</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:center;">状态</th>
        </tr></thead>
        <tbody>${{rows}}
        <tr style="background:rgba(0,212,255,0.08);font-weight:600;">
            <td style="padding:8px 6px;color:#00d4ff;">合计（${{custList.length}}个客户）</td>
            <td style="padding:8px 6px;color:#ff4757;text-align:right;">${{td}}</td>
            <td colspan="6"></td>
        </tr></tbody></table>`;
    document.getElementById('drillModal').classList.add('active');
}}

function renderSellerCustCycle(dept, subDept, seller) {{
    let custList = [];
    if (custBySubDept[dept] && custBySubDept[dept][subDept]) {{
        custList = custBySubDept[dept][subDept].filter(c => c.seller === seller);
    }}
    custList.sort((a, b) => (b.cycle || 0) - (a.cycle || 0));
    const rows = custList.map(c => {{
        const cycle = c.cycle || 0;
        const cycleStr = cycle > 0 ? cycle.toFixed(1) : '-';
        const cc = cycle <= 0 ? '#8892b0' : cycle > 90 ? '#ff4757' : cycle > 60 ? '#ffa502' : '#00ff88';
        const status = cycle <= 0 ? '&#11030; 无数据' : cycle > 90 ? '&#128308; 需关注' : cycle > 60 ? '&#128992; 偏高' : '&#128994; 正常';
        return `<tr style="border-bottom:1px solid rgba(255,255,255,0.05);">
            <td style="padding:7px 5px;color:#ccd6f6;max-width:300px;word-break:break-all;">${{c.customer}}</td>
            <td style="padding:7px 5px;color:#00ff88;text-align:right;">${{(c.collect || 0).toFixed(2)}}</td>
            <td style="padding:7px 5px;color:#ffa502;text-align:right;">${{(c.total_debt || 0).toFixed(2)}}</td>
            <td style="padding:7px 5px;color:${{cc}};text-align:right;font-weight:600;">${{cycleStr}}</td>
            <td style="padding:7px 5px;color:#ccd6f6;text-align:center;">${{status}}</td>
        </tr>`;
    }}).join('');
    const backBtn = `<button onclick="renderSellerCycle('${{dept}}','${{subDept}}')" style="background:rgba(255,255,255,0.1);border:1px solid rgba(255,255,255,0.2);color:#ccd6f6;padding:6px 14px;border-radius:8px;cursor:pointer;font-size:0.85em;margin-bottom:12px;">&larr; 返回销售员列表</button>`;
    document.getElementById('drillModalTitle').innerHTML = `&#9201;&#65039; ${{seller}} - 客户回款周期明细 <span style="font-size:0.7em;color:#8892b0;">（${{subDept}}）</span>`;
    document.getElementById('drillTableContainer').innerHTML = backBtn + `<table style="width:100%;border-collapse:collapse;font-size:0.8em;">
        <thead><tr style="background:rgba(0,212,255,0.12);">
            <th style="padding:8px 6px;color:#00d4ff;text-align:left;">客户名称</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:right;">回款(万)</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:right;">欠款(万)</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:right;">回款周期(天)</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:center;">状态</th>
        </tr></thead>
        <tbody>${{rows}}</tbody></table>`;
    document.getElementById('drillModal').classList.add('active');
}}

function closeDrillModal() {{
    document.getElementById('drillModal').classList.remove('active');
}}

function showOver90Depts() {{
    const over90 = deptData.filter(d => d.cycle > 90);
    if (!over90.length) {{
        alert('当前没有回款周期超过90天的三级部门');
        return;
    }}
    const rows = over90.map(d => `<tr style="border-bottom:1px solid rgba(255,255,255,0.05);cursor:pointer;" onclick="showSellers('${{d.key}}','cycle')">
        <td style="padding:8px 6px;color:#ccd6f6;">&#128194; ${{d.sub_dept}}</td>
        <td style="padding:8px 6px;color:#ff4757;text-align:right;font-weight:600;">${{d.cycle.toFixed(1)}}天</td>
        <td style="padding:8px 6px;color:#ffa502;text-align:right;">${{d.total_debt.toFixed(2)}}万</td>
        <td style="padding:8px 6px;text-align:center;"><span class="status-badge badge-down">需关注</span></td>
    </tr>`).join('');
    document.getElementById('drillModalTitle').textContent = `&#9888;&#65039; 回款周期超90天三级部门（共${{over90.length}}个）`;
    document.getElementById('drillTableContainer').innerHTML = `<table style="width:100%;border-collapse:collapse;font-size:0.85em;">
        <thead><tr style="background:rgba(0,212,255,0.12);">
            <th style="padding:8px 6px;color:#00d4ff;">三级部门</th>
            <th style="padding:8px 6px;color:#ff4757;text-align:right;">回款周期</th>
            <th style="padding:8px 6px;color:#ffa502;text-align:right;">欠款总额</th>
            <th style="padding:8px 6px;color:#00d4ff;text-align:center;">状态</th>
        </tr></thead><tbody>${{rows}}</tbody></table>
        <p style="color:#8892b0;font-size:0.85em;margin-top:10px;text-align:center;">点击部门行可查看销售员回款周期明细</p>`;
    document.getElementById('drillModal').classList.add('active');
}}

Chart.register(ChartDataLabels);
const DEPT_COLORS = ['#00d4ff','#7b2ff7','#00ff88','#ffa502','#ff4757','#ff6b6b','#48dbfb','#ff9f43','#1dd1a1','#5f27cd'];

new Chart(document.getElementById('perfBarChart'), {{
    type: 'bar',
    plugins: [ChartDataLabels],
    data: {{
        labels: {json.dumps(perf_depts, ensure_ascii=False)},
        datasets: [
            {{ label: '25Q2', data: {perf_v25}, backgroundColor: 'rgba(123,47,247,0.5)', borderColor: '#7b2ff7', borderWidth: 1 }},
            {{ label: '26Q2', data: {perf_v26}, backgroundColor: 'rgba(0,212,255,0.7)', borderColor: '#00d4ff', borderWidth: 1 }}
        ]
    }},
    options: {{
        responsive: true, maintainAspectRatio: false,
        plugins: {{ legend: {{ labels: {{ color: '#ccd6f6' }} }}, datalabels: {{ display: false }} }},
        scales: {{
            x: {{ ticks: {{ color: '#8892b0', maxRotation: 45 }}, grid: {{ color: 'rgba(255,255,255,0.05)' }} }},
            y: {{ ticks: {{ color: '#8892b0' }}, grid: {{ color: 'rgba(255,255,255,0.08)' }} }}
        }}
    }}
}});

new Chart(document.getElementById('perfPieChart'), {{
    type: 'doughnut',
    plugins: [ChartDataLabels],
    data: {{
        labels: {json.dumps([x[0] for x in perf_pie], ensure_ascii=False)},
        datasets: [{{ data: {[x[1] for x in perf_pie]}, backgroundColor: DEPT_COLORS, borderWidth: 1, borderColor: '#0a0e27' }}]
    }},
    options: {{
        responsive: true, maintainAspectRatio: false,
        plugins: {{
            legend: {{ position: 'right', labels: {{ color: '#ccd6f6', font: {{ size: 11 }} }} }},
            datalabels: {{
                color: '#fff', font: {{ size: 10 }},
                formatter: (v, ctx) => {{
                    const total = ctx.dataset.data.reduce((a, b) => a + b, 0);
                    return (v/total*100).toFixed(1) + '%';
                }}
            }}
        }}
    }}
}});

new Chart(document.getElementById('debtPieChart'), {{
    type: 'doughnut',
    plugins: [ChartDataLabels],
    data: {{
        labels: {json.dumps(debt_pie_labels, ensure_ascii=False)},
        datasets: [{{ data: {debt_pie_vals}, backgroundColor: ['#00ff88','#ffa502','#ff6b6b','#ff4757'], borderWidth: 1, borderColor: '#0a0e27' }}]
    }},
    options: {{
        responsive: true, maintainAspectRatio: false,
        plugins: {{
            legend: {{ position: 'bottom', labels: {{ color: '#ccd6f6' }} }},
            datalabels: {{
                color: '#fff', font: {{ size: 11, weight: 'bold' }},
                formatter: (v, ctx) => {{
                    const total = ctx.dataset.data.reduce((a, b) => a + b, 0);
                    return (v/total*100).toFixed(1) + '%';
                }}
            }}
        }}
    }}
}});

new Chart(document.getElementById('debtBarChart'), {{
    type: 'bar',
    plugins: [ChartDataLabels],
    data: {{
        labels: {json.dumps(debt_bar_depts, ensure_ascii=False)},
        datasets: [{{ data: {debt_bar_vals}, backgroundColor: DEPT_COLORS, borderWidth: 1, borderColor: '#0a0e27' }}]
    }},
    options: {{
        indexAxis: 'y',
        responsive: true, maintainAspectRatio: false,
        plugins: {{ legend: {{ display: false }}, datalabels: {{ color: '#fff', font: {{ size: 10 }}, anchor: 'end', align: 'end', formatter: v => v.toFixed(1) }} }},
        scales: {{
            x: {{ ticks: {{ color: '#8892b0' }}, grid: {{ color: 'rgba(255,255,255,0.08)' }} }},
            y: {{ ticks: {{ color: '#8892b0', font: {{ size: 11 }} }}, grid: {{ display: false }} }}
        }}
    }}
}});

new Chart(document.getElementById('cycleChart'), {{
    type: 'bar',
    plugins: [ChartDataLabels],
    data: {{
        labels: {json.dumps(chart_cycle_depts, ensure_ascii=False)},
        datasets: [{{
            data: {chart_cycle_vals},
            backgroundColor: {json.dumps(chart_cycle_colors)},
            borderWidth: 1, borderColor: '#0a0e27'
        }}]
    }},
    options: {{
        indexAxis: 'y',
        responsive: true, maintainAspectRatio: false,
        plugins: {{
            legend: {{ display: false }},
            datalabels: {{ color: '#fff', font: {{ size: 11 }}, anchor: 'end', align: 'end', formatter: v => v.toFixed(1) + '天' }},
        }},
        scales: {{
            x: {{ ticks: {{ color: '#8892b0' }}, grid: {{ color: 'rgba(255,255,255,0.08)' }}, max: Math.max(...{chart_cycle_vals}) * 1.15 }},
            y: {{ ticks: {{ color: '#8892b0' }}, grid: {{ display: false }} }}
        }}
    }}
}});
</script>
</body>
</html>'''

outpath = os.path.join(os.path.dirname(os.path.abspath(__file__)), '中西部大区26财年Q2数据看板_弹窗下钻版.html')
with open(outpath, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"\nDone: {outpath}")
print(f"Size: {len(html.encode('utf-8'))/1024:.1f} KB")
