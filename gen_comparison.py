"""整理26Q1 vs 25Q1对比数据：部门→销售员→客户维度
输出Excel文件，包含完成率和同比分析
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import json
from collections import defaultdict, Counter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ========== 常量 ==========
Q1_TARGET = 5586  # Q1目标5586万

INTERNET_SELLERS = ('吴晗', '李国栋', '林茹', '黄佩文', '胡帅帅', '原冀峰')

BLACKLIST = ('支振岗', '李国栋', '白雨')
KEEP_SELLER = ('张宸睿',)

def to_wan(v):
    """转换为万元"""
    if v is None:
        return 0.0
    try:
        return float(v) / 10000
    except (ValueError, TypeError):
        return 0.0

def normalize_dept2(dept2, sub_dept, seller_name):
    """归一化二级部门，只展示2个：湖北营销区、综合管理办公室"""
    dept2 = str(dept2 or '其他').strip().replace('\t', '')
    sub_dept = str(sub_dept or '').strip().replace('\t', '')
    seller = str(seller_name or '').strip().replace('\t', '')

    if dept2 == '武汉仓' or sub_dept == '武汉仓':
        return None

    if seller in INTERNET_SELLERS:
        return '湖北营销区'
    if '通讯互联网' in dept2 or '通讯互联网' in sub_dept:
        return '湖北营销区'

    if dept2 in ('湖北营销区', '综合管理办公室'):
        return dept2

    if dept2 in ('武汉金融', '武汉能源交通', '武汉基建制造', '武汉通讯互联网') or \
       sub_dept in ('武汉金融行业组', '武汉能源交通行业组', '武汉基建制造行业组', '武汉通讯互联网行业组'):
        return '湖北营销区'

    if dept2 in ('中西销售助理部', '华中用户拓展部'):
        return '湖北营销区'

    if dept2 in ('四川营销区', '重庆营销区'):
        return '综合管理办公室'

    if dept2 in ('成都', '重庆', '郑州', '长沙', '西安') or \
       sub_dept in ('成都站', '重庆站', '郑州站', '长沙站', '西安站'):
        return '综合管理办公室'

    if dept2 == '解决方案部':
        return '湖北营销区'

    return '湖北营销区'

def normalize_sub_dept(dept2, sub_dept, raw_dept2='', seller_name=''):
    """归一化三级部门"""
    dept2 = str(dept2 or '').strip().replace('\t', '')
    sub_dept = str(sub_dept or '其他').strip().replace('\t', '')
    raw_dept2 = str(raw_dept2 or '').strip().replace('\t', '')
    seller = str(seller_name or '').strip().replace('\t', '')

    if seller in ('吴晗', '李国栋'):
        return '其他'
    if raw_dept2 in ('中西销售助理部', '华中用户拓展部', '解决方案部'):
        return '其他'
    if '成都' in sub_dept:
        return '成都站'
    if '重庆' in sub_dept:
        return '重庆站'
    return sub_dept

def load_rows_26(filepath):
    """读取26Q1汇总文件"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = {}
        for i, h in enumerate(headers):
            if h:
                r[str(h).strip().replace('\t', '')] = row[i] if i < len(row) else None
        rows.append(r)
    wb.close()
    return rows

def load_rows_25(filepath):
    """读取25Q1汇总文件"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = {}
        for i, h in enumerate(headers):
            if h:
                r[str(h).strip().replace('\t', '')] = row[i] if i < len(row) else None
        rows.append(r)
    wb.close()
    return rows

# ========== 读取数据 ==========
print('读取26Q1数据...')
rows26 = load_rows_26('D:/26财年Q1业绩数据-汇总.xlsx')
print(f'  26Q1总行数: {len(rows26)}')

print('读取25Q1数据...')
rows25 = load_rows_25('D:/25财年Q1业绩数据-汇总.xlsx')
print(f'  25Q1总行数: {len(rows25)}')

# ========== 过滤：只统计核算B端业绩 ==========
perf26 = []
for r in rows26:
    b_flag = str(r.get('是否核算B端业绩', '') or '').strip().replace('\t', '')
    if b_flag != '是':
        continue
    seller = str(r.get('销售员名称', '') or '').strip().replace('\t', '')
    status = str(r.get('销售员状态', '') or '').strip().replace('\t', '')
    raw_dept2 = str(r.get('二级部门', '') or '').strip().replace('\t', '')
    raw_sub_dept = str(r.get('三级部门', '') or '').strip().replace('\t', '')
    dept2 = normalize_dept2(raw_dept2, raw_sub_dept, seller)
    if dept2 is None:
        continue
    sub_dept = normalize_sub_dept(dept2, raw_sub_dept, raw_dept2, seller)
    perf26.append({
        'dept': dept2,
        'sub_dept': sub_dept,
        'seller': seller,
        'customer_id': str(r.get('客户编号', '') or '').strip().replace('\t', ''),
        'customer_name': str(r.get('客户名称', '') or '').strip().replace('\t', ''),
        'perf': to_wan(r.get('业绩总金额')),
        'collect': to_wan(r.get('回款金额')),
        'seller_status': status,
    })

perf25 = []
for r in rows25:
    b_flag = str(r.get('是否核算B端业绩', '') or '').strip().replace('\t', '')
    if b_flag != '是':
        continue
    seller = str(r.get('销售员名称', '') or '').strip().replace('\t', '')
    status = str(r.get('销售员状态', '') or '').strip().replace('\t', '')
    raw_dept2 = str(r.get('二级部门', '') or '').strip().replace('\t', '')
    raw_sub_dept = str(r.get('三级部门', '') or '').strip().replace('\t', '')
    dept2 = normalize_dept2(raw_dept2, raw_sub_dept, seller)
    if dept2 is None:
        continue
    sub_dept = normalize_sub_dept(dept2, raw_sub_dept, raw_dept2, seller)
    perf25.append({
        'dept': dept2,
        'sub_dept': sub_dept,
        'seller': seller,
        'customer_id': str(r.get('客户编号', '') or '').strip().replace('\t', ''),
        'customer_name': str(r.get('客户名称', '') or '').strip().replace('\t', ''),
        'perf': to_wan(r.get('业绩总金额')),
        'collect': to_wan(r.get('回款金额')),
        'seller_status': status,
    })

print(f'  26Q1核算B端行数: {len(perf26)}')
print(f'  25Q1核算B端行数: {len(perf25)}')

# ========== 在职销售员口径 ==========
# 26Q1：有核算B端业绩（>0）的在职销售员，剔除黑名单，保留张宸睿
seller_perf26 = defaultdict(float)
for r in perf26:
    seller_perf26[r['seller']] += r['perf']

active_sellers_26 = set()
for s, p in seller_perf26.items():
    if s in BLACKLIST and s not in KEEP_SELLER:
        continue
    if p > 0:
        status_list = [r['seller_status'] for r in perf26 if r['seller'] == s]
        if any(st == '在职' for st in status_list):
            active_sellers_26.add(s)

# 25Q1同理
seller_perf25 = defaultdict(float)
for r in perf25:
    seller_perf25[r['seller']] += r['perf']

active_sellers_25 = set()
for s, p in seller_perf25.items():
    if s in BLACKLIST and s not in KEEP_SELLER:
        continue
    if p > 0:
        status_list = [r['seller_status'] for r in perf25 if r['seller'] == s]
        if any(st == '在职' for st in status_list):
            active_sellers_25.add(s)

# ========== 聚合维度 ==========

# 1. 部门维度
dept_agg26 = defaultdict(lambda: {'perf': 0.0, 'collect': 0.0, 'sellers': set()})
dept_agg25 = defaultdict(lambda: {'perf': 0.0, 'collect': 0.0, 'sellers': set()})

for r in perf26:
    d = r['dept']
    dept_agg26[d]['perf'] += r['perf']
    dept_agg26[d]['collect'] += r['collect']
    if r['seller'] in active_sellers_26:
        dept_agg26[d]['sellers'].add(r['seller'])

for r in perf25:
    d = r['dept']
    dept_agg25[d]['perf'] += r['perf']
    dept_agg25[d]['collect'] += r['collect']
    if r['seller'] in active_sellers_25:
        dept_agg25[d]['sellers'].add(r['seller'])

# 1.5 二级+三级部门维度
subdept_agg26 = defaultdict(lambda: {'perf': 0.0, 'collect': 0.0, 'sellers': set()})
subdept_agg25 = defaultdict(lambda: {'perf': 0.0, 'collect': 0.0, 'sellers': set()})

for r in perf26:
    key = (r['dept'], r['sub_dept'])
    subdept_agg26[key]['perf'] += r['perf']
    subdept_agg26[key]['collect'] += r['collect']
    if r['seller'] in active_sellers_26:
        subdept_agg26[key]['sellers'].add(r['seller'])

for r in perf25:
    key = (r['dept'], r['sub_dept'])
    subdept_agg25[key]['perf'] += r['perf']
    subdept_agg25[key]['collect'] += r['collect']
    if r['seller'] in active_sellers_25:
        subdept_agg25[key]['sellers'].add(r['seller'])

# 2. 销售员维度
seller_agg26 = defaultdict(lambda: {'perf': 0.0, 'collect': 0.0, 'dept': '', 'sub_dept': ''})
seller_agg25 = defaultdict(lambda: {'perf': 0.0, 'collect': 0.0, 'dept': '', 'sub_dept': ''})

for r in perf26:
    s = r['seller']
    seller_agg26[s]['perf'] += r['perf']
    seller_agg26[s]['collect'] += r['collect']
    seller_agg26[s]['dept'] = r['dept']
    seller_agg26[s]['sub_dept'] = r['sub_dept']

for r in perf25:
    s = r['seller']
    seller_agg25[s]['perf'] += r['perf']
    seller_agg25[s]['collect'] += r['collect']
    seller_agg25[s]['dept'] = r['dept']
    seller_agg25[s]['sub_dept'] = r['sub_dept']

# 3. 客户维度（按部门+销售员+客户聚合）
cust_agg26 = defaultdict(lambda: {'perf': 0.0, 'collect': 0.0, 'dept': '', 'sub_dept': '', 'seller': ''})
cust_agg25 = defaultdict(lambda: {'perf': 0.0, 'collect': 0.0, 'dept': '', 'sub_dept': '', 'seller': ''})

for r in perf26:
    key = (r['dept'], r['seller'], r['customer_id'])
    cust_agg26[key]['perf'] += r['perf']
    cust_agg26[key]['collect'] += r['collect']
    cust_agg26[key]['dept'] = r['dept']
    cust_agg26[key]['sub_dept'] = r['sub_dept']
    cust_agg26[key]['seller'] = r['seller']
    cust_agg26[key]['customer_id'] = r['customer_id']
    cust_agg26[key]['customer_name'] = r['customer_name']

for r in perf25:
    key = (r['dept'], r['seller'], r['customer_id'])
    cust_agg25[key]['perf'] += r['perf']
    cust_agg25[key]['collect'] += r['collect']
    cust_agg25[key]['dept'] = r['dept']
    cust_agg25[key]['sub_dept'] = r['sub_dept']
    cust_agg25[key]['seller'] = r['seller']
    cust_agg25[key]['customer_id'] = r['customer_id']
    cust_agg25[key]['customer_name'] = r['customer_name']

# ========== 总计 ==========
total_perf26 = sum(d['perf'] for d in dept_agg26.values())
total_perf25 = sum(d['perf'] for d in dept_agg25.values())
total_collect26 = sum(d['collect'] for d in dept_agg26.values())
total_collect25 = sum(d['collect'] for d in dept_agg25.values())

yoy = (total_perf26 - total_perf25) / total_perf25 * 100 if total_perf25 else 0
completion_rate = total_perf26 / Q1_TARGET * 100

print(f'26Q1总业绩: {total_perf26:.2f}万')
print(f'25Q1总业绩: {total_perf25:.2f}万')
print(f'同比: {yoy:.1f}%')
print(f'完成率: {completion_rate:.1f}%')
print(f'在职销售员26Q1: {len(active_sellers_26)}人')
print(f'在职销售员25Q1: {len(active_sellers_25)}人')

# ========== 生成Excel ==========
wb_out = Workbook()

# 样式定义
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', start_color='1F4E79')
header_align = Alignment(horizontal='center', vertical='center')

title_font = Font(name='Arial', bold=True, size=14, color='1F4E79')
subtitle_font = Font(name='Arial', bold=True, size=12, color='2E75B6')

data_font = Font(name='Arial', size=10)
number_font = Font(name='Arial', size=10)
red_font = Font(name='Arial', size=10, color='FF0000')
green_font = Font(name='Arial', size=10, color='008000')
bold_font = Font(name='Arial', size=10, bold=True)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

highlight_fill = PatternFill('solid', start_color='D6E4F0')
total_fill = PatternFill('solid', start_color='BDD7EE')

def fmt_pct(v):
    """格式化百分比"""
    if v is None or v == 0:
        return '-'
    return f'{v:.1f}%'

def fmt_wan(v):
    """格式化万元"""
    if v == 0:
        return '0.00'
    return f'{v:.2f}'

def write_header_row(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

def write_data_row(ws, row, values, is_total=False, is_highlight=False):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = bold_font if is_total else data_font
        cell.border = thin_border
        if is_total:
            cell.fill = total_fill
        elif is_highlight:
            cell.fill = highlight_fill
        # 数值类型右对齐
        if isinstance(v, (int, float)):
            cell.alignment = Alignment(horizontal='right')
        elif isinstance(v, str) and '%' in str(v):
            cell.alignment = Alignment(horizontal='right')
            # 同比红色/绿色
            if v.startswith('-'):
                cell.font = red_font if not is_total else Font(name='Arial', size=10, bold=True, color='FF0000')
            elif v != '-' and not v.startswith('0'):
                cell.font = green_font if not is_total else Font(name='Arial', size=10, bold=True, color='008000')

# ========== Sheet1: 总览 ==========
ws1 = wb_out.active
ws1.title = '总览'

ws1.cell(row=1, column=1, value='26财年Q1 vs 25财年Q1 基础完成情况对比').font = title_font
ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

ws1.cell(row=2, column=1, value='口径：只统计是否核算B端业绩=是；部门合并后只展示湖北营销区、综合管理办公室').font = subtitle_font
ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

headers_total = ['指标', '26Q1', '25Q1', '同比', 'Q1目标', '完成率', '在职销售员26Q1', '在职销售员25Q1']
write_header_row(ws1, 4, headers_total)

yoy_str = fmt_pct(yoy)
cr_str = fmt_pct(completion_rate)

write_data_row(ws1, 5, ['总业绩(万)', fmt_wan(total_perf26), fmt_wan(total_perf25), yoy_str, Q1_TARGET, cr_str, len(active_sellers_26), len(active_sellers_25)], is_total=True)

# 部门行
row_idx = 6
for dept in ['湖北营销区', '综合管理办公室']:
    d26 = dept_agg26.get(dept, {'perf': 0, 'collect': 0, 'sellers': set()})
    d25 = dept_agg25.get(dept, {'perf': 0, 'collect': 0, 'sellers': set()})
    perf26_d = d26['perf']
    perf25_d = d25['perf']
    yoy_d = (perf26_d - perf25_d) / perf25_d * 100 if perf25_d else None
    cr_d = perf26_d / Q1_TARGET * 100  # 部门完成率是按总目标还是部门目标？按总目标
    sellers26_d = len(d26['sellers'])
    sellers25_d = len(d25['sellers'])
    write_data_row(ws1, row_idx, [dept, fmt_wan(perf26_d), fmt_wan(perf25_d), fmt_pct(yoy_d), Q1_TARGET, fmt_pct(cr_d), sellers26_d, sellers25_d])
    row_idx += 1

# 列宽
ws1.column_dimensions['A'].width = 18
ws1.column_dimensions['B'].width = 14
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 12
ws1.column_dimensions['G'].width = 18
ws1.column_dimensions['H'].width = 18

# ========== Sheet2: 部门维度 ==========
ws2 = wb_out.create_sheet('部门维度')

ws2.cell(row=1, column=1, value='部门维度：26Q1 vs 25Q1业绩对比').font = title_font
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

headers_dept = ['二级部门', '26Q1业绩(万)', '26Q1回款(万)', '25Q1业绩(万)', '25Q1回款(万)', '同比(%)', 'Q1目标(万)', '完成率(%)', '在职销售员26Q1', '在职销售员25Q1']
write_header_row(ws2, 3, headers_dept)

row_idx = 4
for dept in ['湖北营销区', '综合管理办公室']:
    d26 = dept_agg26.get(dept, {'perf': 0, 'collect': 0, 'sellers': set()})
    d25 = dept_agg25.get(dept, {'perf': 0, 'collect': 0, 'sellers': set()})
    perf26_d = d26['perf']
    collect26_d = d26['collect']
    perf25_d = d25['perf']
    collect25_d = d25['collect']
    yoy_d = (perf26_d - perf25_d) / perf25_d * 100 if perf25_d else None
    cr_d = perf26_d / Q1_TARGET * 100
    sellers26_d = len(d26['sellers'])
    sellers25_d = len(d25['sellers'])
    write_data_row(ws2, row_idx, [dept, perf26_d, collect26_d, perf25_d, collect25_d, fmt_pct(yoy_d), Q1_TARGET, fmt_pct(cr_d), sellers26_d, sellers25_d])
    row_idx += 1

# 合计行
write_data_row(ws2, row_idx, ['合计', total_perf26, total_collect26, total_perf25, total_collect25, fmt_pct(yoy), Q1_TARGET, fmt_pct(completion_rate), len(active_sellers_26), len(active_sellers_25)], is_total=True)

for col_letter in ['A','B','C','D','E','F','G','H','I','J']:
    ws2.column_dimensions[col_letter].width = 16

# ========== Sheet2.5: 部门-三级部门维度 ==========
ws25 = wb_out.create_sheet('部门-三级部门维度')

ws25.cell(row=1, column=1, value='三级部门维度：26Q1 vs 25Q1业绩对比').font = title_font
ws25.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

ws25.cell(row=2, column=1, value='口径：只统计是否核算B端业绩=是；三级部门按映射规则归一化').font = subtitle_font
ws25.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)

headers_subdept = ['二级部门', '三级部门', '26Q1业绩(万)', '26Q1回款(万)', '25Q1业绩(万)', '25Q1回款(万)', '同比(%)', 'Q1目标(万)', '完成率(%)', '在职销售员26Q1', '在职销售员25Q1']
write_header_row(ws25, 4, headers_subdept)

# 按二级部门分组，内部按26Q1业绩降序排列
dept_order = ['湖北营销区', '综合管理办公室']
all_subdept_keys = set(subdept_agg26.keys()) | set(subdept_agg25.keys())

row_idx = 5
for dept in dept_order:
    # 找到该部门下的所有三级部门
    dept_subdepts = sorted(
        [k for k in all_subdept_keys if k[0] == dept],
        key=lambda k: -(subdept_agg26.get(k, {}).get('perf', 0) or 0)
    )
    for key in dept_subdepts:
        sd26 = subdept_agg26.get(key, {'perf': 0, 'collect': 0, 'sellers': set()})
        sd25 = subdept_agg25.get(key, {'perf': 0, 'collect': 0, 'sellers': set()})
        perf26_sd = sd26['perf']
        collect26_sd = sd26['collect']
        perf25_sd = sd25['perf']
        collect25_sd = sd25['collect']
        yoy_sd = (perf26_sd - perf25_sd) / perf25_sd * 100 if perf25_sd else None
        cr_sd = perf26_sd / Q1_TARGET * 100
        sellers26_sd = len(sd26['sellers'])
        sellers25_sd = len(sd25['sellers'])
        write_data_row(ws25, row_idx, [dept, key[1], perf26_sd, collect26_sd, perf25_sd, collect25_sd, fmt_pct(yoy_sd), Q1_TARGET, fmt_pct(cr_sd), sellers26_sd, sellers25_sd])
        row_idx += 1
    # 部门小计
    d26 = dept_agg26.get(dept, {'perf': 0, 'collect': 0, 'sellers': set()})
    d25 = dept_agg25.get(dept, {'perf': 0, 'collect': 0, 'sellers': set()})
    yoy_dept = (d26['perf'] - d25['perf']) / d25['perf'] * 100 if d25['perf'] else None
    cr_dept = d26['perf'] / Q1_TARGET * 100
    write_data_row(ws25, row_idx, [dept + ' 小计', '', d26['perf'], d26['collect'], d25['perf'], d25['collect'], fmt_pct(yoy_dept), Q1_TARGET, fmt_pct(cr_dept), len(d26['sellers']), len(d25['sellers'])], is_highlight=True)
    row_idx += 1

# 总合计
write_data_row(ws25, row_idx, ['合计', '', total_perf26, total_collect26, total_perf25, total_collect25, fmt_pct(yoy), Q1_TARGET, fmt_pct(completion_rate), len(active_sellers_26), len(active_sellers_25)], is_total=True)

for col_letter in ['A','B','C','D','E','F','G','H','I','J','K']:
    ws25.column_dimensions[col_letter].width = 16
ws25.column_dimensions['A'].width = 18
ws25.column_dimensions['B'].width = 22

# ========== Sheet3: 销售员维度 ==========
ws3 = wb_out.create_sheet('销售员维度')

ws3.cell(row=1, column=1, value='销售员维度：26Q1 vs 25Q1业绩对比（在职销售员）').font = title_font
ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

headers_seller = ['二级部门', '三级部门', '销售员', '26Q1业绩(万)', '26Q1回款(万)', '25Q1业绩(万)', '25Q1回款(万)', '同比(%)', '26Q1完成率(%)', '人均26Q1(万)', '人均25Q1(万)']
write_header_row(ws3, 3, headers_seller)

# 合并所有在职销售员（26Q1在职且26Q1有业绩）
all_active_sellers = sorted(active_sellers_26, key=lambda s: (
    seller_agg26.get(s, {}).get('dept', ''),
    seller_agg26.get(s, {}).get('sub_dept', ''),
    -seller_agg26.get(s, {}).get('perf', 0)
))

row_idx = 4
dept_total_26 = 0
dept_total_25 = 0
prev_dept = None

for s in all_active_sellers:
    s26 = seller_agg26.get(s, {'perf': 0, 'collect': 0, 'dept': '', 'sub_dept': ''})
    s25 = seller_agg25.get(s, {'perf': 0, 'collect': 0, 'dept': '', 'sub_dept': ''})
    perf26_s = s26['perf']
    collect26_s = s26['collect']
    perf25_s = s25['perf']
    collect25_s = s25['collect']
    dept_s = s26['dept'] if s26['dept'] else s25['dept']
    sub_dept_s = s26['sub_dept'] if s26['sub_dept'] else s25['sub_dept']
    yoy_s = (perf26_s - perf25_s) / perf25_s * 100 if perf25_s else None
    cr_s = perf26_s / Q1_TARGET * 100

    write_data_row(ws3, row_idx, [dept_s, sub_dept_s, s, perf26_s, collect26_s, perf25_s, collect25_s, fmt_pct(yoy_s), fmt_pct(cr_s), '', ''])
    row_idx += 1

# 合计行
avg26 = total_perf26 / len(active_sellers_26) if active_sellers_26 else 0
avg25 = total_perf25 / len(active_sellers_25) if active_sellers_25 else 0
write_data_row(ws3, row_idx, ['合计', '', f'{len(active_sellers_26)}人', total_perf26, total_collect26, total_perf25, total_collect25, fmt_pct(yoy), fmt_pct(completion_rate), avg26, avg25], is_total=True)

for col_letter in ['A','B','C','D','E','F','G','H','I','J','K']:
    ws3.column_dimensions[col_letter].width = 16
ws3.column_dimensions['C'].width = 12

# ========== Sheet4: 客户维度 ==========
ws4 = wb_out.create_sheet('客户维度')

ws4.cell(row=1, column=1, value='客户维度：26Q1 vs 25Q1业绩对比').font = title_font
ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

headers_cust = ['二级部门', '三级部门', '销售员', '客户编号', '客户名称', '26Q1业绩(万)', '26Q1回款(万)', '25Q1业绩(万)', '25Q1回款(万)', '同比(%)']
write_header_row(ws4, 3, headers_cust)

# 按26Q1有业绩的客户排序，也包含25Q1有但26Q1没有的客户
all_cust_keys = set(cust_agg26.keys()) | set(cust_agg25.keys())
# 按部门、销售员、26Q1业绩排序
sorted_custs = sorted(all_cust_keys, key=lambda k: (
    cust_agg26.get(k, cust_agg25.get(k, {})).get('dept', ''),
    cust_agg26.get(k, cust_agg25.get(k, {})).get('seller', ''),
    -(cust_agg26.get(k, {}).get('perf', 0) or 0)
))

row_idx = 4
for key in sorted_custs:
    c26 = cust_agg26.get(key, {'perf': 0, 'collect': 0, 'dept': '', 'sub_dept': '', 'seller': '', 'customer_id': '', 'customer_name': ''})
    c25 = cust_agg25.get(key, {'perf': 0, 'collect': 0, 'dept': '', 'sub_dept': '', 'seller': '', 'customer_id': '', 'customer_name': ''})

    dept_c = c26['dept'] if c26['dept'] else c25['dept']
    sub_dept_c = c26['sub_dept'] if c26['sub_dept'] else c25['sub_dept']
    seller_c = c26['seller'] if c26['seller'] else c25['seller']
    cust_id = c26.get('customer_id', '') or c25.get('customer_id', '')
    cust_name = c26.get('customer_name', '') or c25.get('customer_name', '')

    perf26_c = c26['perf']
    perf25_c = c25['perf']
    collect26_c = c26['collect']
    collect25_c = c25['collect']
    yoy_c = (perf26_c - perf25_c) / perf25_c * 100 if perf25_c else None

    write_data_row(ws4, row_idx, [dept_c, sub_dept_c, seller_c, cust_id, cust_name, perf26_c, collect26_c, perf25_c, collect25_c, fmt_pct(yoy_c)])
    row_idx += 1

for col_letter in ['A','B','C','D','E','F','G','H','I','J']:
    ws4.column_dimensions[col_letter].width = 16
ws4.column_dimensions['E'].width = 30

# ========== 保存 ==========
output_path = 'C:/Users/wm881/WorkBuddy/2026-06-09-17-16-53/zhongxibu-dashboard/26Q1_vs_25Q1_对比分析.xlsx'
wb_out.save(output_path)
print(f'已保存到: {output_path}')
print(f'Sheet1: 总览 - KPI概览')
print(f'Sheet2: 部门维度 - 2个部门对比')
print(f'Sheet3: 销售员维度 - {len(all_active_sellers)}个在职销售员对比')
print(f'Sheet4: 客户维度 - {len(sorted_custs)}个客户对比')
