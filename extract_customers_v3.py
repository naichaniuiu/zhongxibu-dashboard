import openpyxl
import json
import os
from collections import defaultdict
from datetime import datetime

# ============================================================
# 中西部大区客户明细提取 - 新数据源
# ============================================================

TODAY = datetime(2026, 7, 1)  # 数据基准日（数据拉取日）
DATA_FILE = 'D:/业绩欠款看板.xlsx'  # 单文件多Sheet数据源
Q1_START = datetime(2026, 4, 1)
Q1_END = datetime(2026, 6, 30)

DEPT_MAP = {
    '中西部大区': '中西部大区',
    '中西大区': '中西部大区',
    '华中大区': '中西部大区',
    '华中大区（已封存）': '中西部大区',
    '西南大区': '中西部大区',
    '西南大区（已封存）': '中西部大区',
}


# ============================================================
# 部门映射规则（二级部门 & 三级部门）
# ============================================================
INTERNET_SELLERS = set()


def scan_internet_sellers(path, sheet_idx=0):
    """扫描文件中原始部门为武汉通讯互联网的销售员，用于后续按销售员归属映射。"""
    for r in load_rows(path, sheet_idx=sheet_idx):
        dept1 = str(r.get('一级部门') or '').strip().replace('\t', '')
        if dept1 not in DEPT_MAP:
            continue
        dept2 = str(r.get('二级部门') or '').strip().replace('\t', '')
        sub_dept = str(r.get('三级部门') or '').strip().replace('\t', '')
        if '通讯互联网' in dept2 or '通讯互联网' in sub_dept:
            seller = str(r.get('销售员名称') or '').strip().replace('\t', '')
            if seller:
                INTERNET_SELLERS.add(seller)


def normalize_dept2(dept2, sub_dept, seller_name):
    """根据二级部门、三级部门和销售员名称归一化二级部门。
    
    目标只展示 2 个部门：湖北营销区、综合管理办公室。
    混营销区已合并为综合管理办公室；武汉仓 → 删除（返回 None）。
    """
    dept2 = str(dept2 or '其他').strip().replace('\t', '')
    sub_dept = str(sub_dept or '').strip().replace('\t', '')
    seller = str(seller_name or '').strip().replace('\t', '')

    # 0. 武汉仓 → 删除，不纳入任何部门
    if dept2 == '武汉仓' or sub_dept == '武汉仓':
        return None

    # 1. 武汉通讯互联网：所有销售员 → 湖北营销区（吴晗/李国栋已改为湖北营销区）
    if seller in INTERNET_SELLERS:
        return '湖北营销区'
    if '通讯互联网' in dept2 or '通讯互联网' in sub_dept:
        return '湖北营销区'

    # 2. 已经是目标部门名的，直接返回
    if dept2 in ('湖北营销区', '综合管理办公室'):
        return dept2

    # 3. 武汉金融 / 武汉能源交通 / 武汉基建制造 / 武汉通讯互联网 → 湖北营销区
    if dept2 in ('武汉金融', '武汉能源交通', '武汉基建制造', '武汉通讯互联网') or \
       sub_dept in ('武汉金融行业组', '武汉能源交通行业组', '武汉基建制造行业组', '武汉通讯互联网行业组'):
        return '湖北营销区'

    # 4. 中西销售助理部 / 华中用户拓展部 → 湖北营销区
    if dept2 in ('中西销售助理部', '华中用户拓展部'):
        return '湖北营销区'

    # 5. 四川营销区 / 重庆营销区 → 综合管理办公室
    if dept2 in ('四川营销区', '重庆营销区'):
        return '综合管理办公室'

    # 6. 其他城市部门：成都 / 重庆 / 郑州 / 长沙 / 西安 → 综合管理办公室
    if dept2 in ('成都', '重庆', '郑州', '长沙', '西安') or \
       sub_dept in ('成都站', '重庆站', '郑州站', '长沙站', '西安站'):
        return '综合管理办公室'

    # 7. 解决方案部 → 湖北营销区
    if dept2 == '解决方案部':
        return '湖北营销区'

    # 8. 兜底：其他未识别的部门 → 湖北营销区
    return '湖北营销区'


def normalize_sub_dept(dept2, sub_dept, raw_dept2='', seller_name=''):
    """根据归一化后的二级部门、原始二级部门和销售员名称，归一化三级部门。"""
    dept2 = str(dept2 or '').strip().replace('\t', '')
    sub_dept = str(sub_dept or '其他').strip().replace('\t', '')
    raw_dept2 = str(raw_dept2 or '').strip().replace('\t', '')
    seller = str(seller_name or '').strip().replace('\t', '')

    # 吴晗/李国栋归入湖北营销区后，三级统一为其他
    if seller in ('吴晗', '李国栋'):
        return '其他'

    # 中西销售助理部/华中用户拓展部/解决方案部归入湖北营销区后，三级统一为其他
    if raw_dept2 in ('中西销售助理部', '华中用户拓展部', '解决方案部'):
        return '其他'

    if '成都' in sub_dept:
        return '成都站'
    if '重庆' in sub_dept:
        return '重庆站'

    return sub_dept

def parse_date(val):
    if not val:
        return None
    s = str(val).strip().replace('\t', '').split(' ')[0]
    if not s or s == 'None':
        return None
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None

def to_wan(val):
    if val is None:
        return 0.0
    try:
        return float(val) / 10000.0
    except (TypeError, ValueError):
        return 0.0

def load_rows(path, sheet_idx=0):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[sheet_idx]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield dict(zip(headers, row))

# 聚合结构：dept -> seller -> customer -> dict
result = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {
    'customer': '',
    'perf': 0.0,
    'collect': 0.0,
    'total_debt': 0.0,
    'd30': 0.0,
    'd30_90': 0.0,
    'd90_180': 0.0,
    'd180': 0.0,
    'max_days': 0,
    'cycle': 0.0,
    'cycle_weight': 0.0,
    'cycle_days_weighted': 0.0,
    'orders': 0,
})))

# 0. 预扫描：识别武汉通讯互联网部门的销售员
print('Scanning internet sales sellers...')
scan_internet_sellers(DATA_FILE, sheet_idx=1)  # 26Q1业绩
scan_internet_sellers(DATA_FILE, sheet_idx=2)  # 欠款数据
print(f'  Internet sellers: {len(INTERNET_SELLERS)}')

# 1. 业绩数据：Q1 中西部大区
print('Processing performance data...')
for r in load_rows(DATA_FILE, sheet_idx=1):
    dept1 = str(r.get('一级部门') or '').strip().replace('\t', '')
    if dept1 not in DEPT_MAP:
        continue
    d = parse_date(r.get('业绩日期'))
    if not d or not (Q1_START <= d <= Q1_END):
        continue
    is_b = str(r.get('是否核算B端业绩') or '').strip().replace('\t', '')
    if is_b == '否':
        continue
    raw_dept2 = str(r.get('二级部门') or '').strip().replace('\t', '')
    raw_sub_dept = str(r.get('三级部门') or '').strip().replace('\t', '')
    seller = str(r.get('销售员名称') or '').strip().replace('\t', '')
    customer = str(r.get('客户名称') or '').strip().replace('\t', '') or '未知客户'
    dept = normalize_dept2(raw_dept2, raw_sub_dept, seller)
    if dept is None:
        continue  # 武汉仓等删除部门，跳过
    sub_dept = normalize_sub_dept(dept, raw_sub_dept, raw_dept2, seller)
    perf = to_wan(r.get('业绩总金额'))
    collect = to_wan(r.get('回款金额'))
    debt = to_wan(r.get('欠款金额'))
    
    c = result[dept][seller][customer]
    c['customer'] = customer
    c['sub_dept'] = sub_dept
    c['perf'] += perf
    c['collect'] += collect
    c['total_debt'] += debt
    c['orders'] += 1

# 2. 欠款数据：整个文件所有行直接求和，无过滤
print('Processing debt data...')
for r in load_rows(DATA_FILE, sheet_idx=2):
    d = parse_date(r.get('业绩日期'))
    raw_dept2 = str(r.get('二级部门') or '').strip().replace('\t', '')
    raw_sub_dept = str(r.get('三级部门') or '').strip().replace('\t', '')
    seller = str(r.get('销售员名称') or '').strip().replace('\t', '')
    customer = str(r.get('客户名称') or '').strip().replace('\t', '') or '未知客户'
    dept = normalize_dept2(raw_dept2, raw_sub_dept, seller)
    if dept is None:
        continue  # 武汉仓等删除部门，跳过
    debt_val = to_wan(r.get('欠款金额'))
    days = (TODAY - d).days if d else 0
    if days < 0:
        days = 0
    
    c = result[dept][seller][customer]
    c['customer'] = customer
    c['total_debt'] += debt_val
    c['orders'] += 1
    if days > c['max_days']:
        c['max_days'] = days
    
    if days <= 30:
        c['d30'] += debt_val
    elif days <= 90:
        c['d30_90'] += debt_val
    elif days <= 180:
        c['d90_180'] += debt_val
    else:
        c['d180'] += debt_val

# 3. 认款数据：计算客户回款周期
print('Processing payment data...')
order_to_dept = {}
for dept, sellers in result.items():
    for seller, customers in sellers.items():
        for customer, data in customers.items():
            if data['perf'] > 0:
                order_to_dept.setdefault(customer, dept)

for r in load_rows(DATA_FILE, sheet_idx=3):
    if str(r.get('目标认款类型') or '').strip() != '业绩单认款':
        continue
    perf_date = parse_date(r.get('业绩日期'))
    pay_date = parse_date(r.get('回款日期'))
    if not perf_date or not pay_date:
        continue
    days = (pay_date - perf_date).days
    if days < 0 or days > 365 * 2:
        continue
    amount = to_wan(r.get('认款协同金额'))
    seller = str(r.get('销售员名称') or '').strip().replace('\t', '')
    customer = str(r.get('回款客户名称') or '').strip().replace('\t', '') or '未知客户'
    # 通过销售员+客户定位到部门
    dept = None
    for d in result.keys():
        if seller in result[d] and customer in result[d][seller]:
            dept = d
            break
    if not dept:
        continue
    
    c = result[dept][seller][customer]
    c['cycle_weight'] += amount
    c['cycle_days_weighted'] += days * amount

# 整理输出
print('Building customer_detail.json...')
output = {}
for dept, sellers in result.items():
    output[dept] = {}
    for seller, customers in sellers.items():
        customers_list = []
        for customer, data in customers.items():
            if data['cycle_weight'] > 0:
                data['cycle'] = round(data['cycle_days_weighted'] / data['cycle_weight'], 1)
            else:
                data['cycle'] = 0.0
            del data['cycle_weight']
            del data['cycle_days_weighted']
            # 如果客户没有任何业绩/欠款/回款，跳过
            if data['perf'] <= 0 and data['total_debt'] <= 0 and data['collect'] <= 0:
                continue
            customers_list.append({
                'customer': data['customer'],
                'sub_dept': data.get('sub_dept', '其他') or '其他',
                'perf': round(data['perf'], 2),
                'collect': round(data['collect'], 2),
                'total_debt': round(data['total_debt'], 2),
                'd30': round(data['d30'], 2),
                'd30_90': round(data['d30_90'], 2),
                'd90_180': round(data['d90_180'], 2),
                'd180': round(data['d180'], 2),
                'max_days': data['max_days'],
                'cycle': data['cycle'],
                'orders': data['orders'],
            })
        # 按欠款金额降序排列
        customers_list.sort(key=lambda x: -x['total_debt'])
        if customers_list:
            output[dept][seller] = customers_list

# 只保留有数据的部门
output = {k: v for k, v in output.items() if v}

base_dir = os.path.dirname(os.path.abspath(__file__))
with open(os.path.join(base_dir, 'customer_detail.json'), 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print('Done.')
print(f'Departments: {len(output)}')
for dept, sellers in output.items():
    print(f'  {dept}: {len(sellers)} sellers, {sum(len(v) for v in sellers.values())} customers')
